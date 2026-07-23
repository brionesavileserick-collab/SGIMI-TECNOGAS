"""
Reports service layer – Generate, export, and persist reports.

Expansiones implementadas
─────────────────────────
 1  Filtros por fecha       – date_from / date_to en todos los métodos
 2  Filtro por producto     – product_id en inventario, movimientos, discrepancias
 3  Filtro por usuario      – user_id en actividad y movimientos
 4  Exportación CSV         – export_to_csv(report_data, filename)
 5  Exportación Excel       – export_to_excel(report_data, filename)
 6  Vista formateada        – format_report_for_display(report_type, report_data)
 7  Reportes guardados      – save / list / get / delete report configs
 8  Comparación de períodos – generate_comparison_report(...)
 9  Top productos           – generate_top_products_report(...)
10  Eficiencia por sucursal – generate_branch_efficiency_report(...)
11  Valor de inventario     – generate_inventory_value_report(...)
12  Reporte de transferencias – generate_transfer_report(...)
13  Tendencias              – generate_trend_report(...)

Principios
──────────
• Este módulo SOLO lee datos, nunca escribe/modifica registros de negocio.
• Todos los parámetros nuevos son opcionales (default None / 10 / etc.).
• Los métodos originales mantienen su firma extendida de forma retrocompatible.
"""

import csv
import io
import json
import logging
import os
import re
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional

from sqlalchemy import Boolean, Column, DateTime, ForeignKey, Integer, String, Text, func, or_
from sqlalchemy.orm import Session

from core.database import Base

logger = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════════════════════════
# Helpers
# ══════════════════════════════════════════════════════════════════════════════

def _fmt_date(dt: Optional[datetime]) -> Optional[str]:
    """ISO-format a datetime or return None."""
    return dt.isoformat() if dt else None


def _flatten_dict(d: Dict, parent_key: str = "", sep: str = ".") -> Dict:
    """Recursively flatten a nested dict for CSV export."""
    items: List = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(_flatten_dict(v, new_key, sep=sep).items())
        elif isinstance(v, list):
            items.append((new_key, json.dumps(v, ensure_ascii=False)))
        else:
            items.append((new_key, v))
    return dict(items)


def _fmt_date_readable(dt: Optional[datetime]) -> Optional[str]:
    """Format datetime to readable format (DD/MM/YYYY HH:MM)."""
    if dt is None:
        return None
    return dt.strftime("%d/%m/%Y %H:%M")


def _fmt_date_iso_to_readable(iso_date: Optional[str]) -> str:
    """Convert ISO date string to readable format."""
    if not iso_date or iso_date == "—":
        return "—"
    try:
        dt = datetime.fromisoformat(iso_date.replace("Z", "+00:00"))
        return _fmt_date_readable(dt)
    except Exception:
        return iso_date


# ══════════════════════════════════════════════════════════════════════════════
# Service class
# ══════════════════════════════════════════════════════════════════════════════

class ReportSchedule(Base):
    """Simple persisted definition for recurring report schedules."""

    __tablename__ = "report_schedules"

    id = Column(Integer, primary_key=True, index=True)
    name = Column(String(100), nullable=False)
    report_type = Column(String(50), nullable=False)
    schedule_type = Column(String(20), nullable=False)
    schedule_time = Column(String(5), nullable=False)
    schedule_day_of_week = Column(Integer, nullable=True)
    schedule_day_of_month = Column(Integer, nullable=True)
    parameters = Column(Text, nullable=True)
    last_run_at = Column(DateTime(timezone=True), nullable=True)
    next_run_at = Column(DateTime(timezone=True), nullable=True)
    is_active = Column(Boolean, default=True, nullable=False)
    created_by_user_id = Column(Integer, nullable=True)
    created_at = Column(DateTime(timezone=True), server_default=func.now())

    def to_dict(self) -> Dict[str, Any]:
        return {
            "id": self.id,
            "name": self.name,
            "report_type": self.report_type,
            "schedule_type": self.schedule_type,
            "schedule_time": self.schedule_time,
            "schedule_day_of_week": self.schedule_day_of_week,
            "schedule_day_of_month": self.schedule_day_of_month,
            "parameters": json.loads(self.parameters or "{}") if self.parameters else {},
            "last_run_at": _fmt_date(self.last_run_at),
            "next_run_at": _fmt_date(self.next_run_at),
            "is_active": self.is_active,
            "created_by_user_id": self.created_by_user_id,
            "created_at": _fmt_date(self.created_at),
        }


class ReportsService:
    """Service for generating, exporting, and persisting reports."""

    def __init__(self, db: Session):
        self.db = db
        self._ensure_schedule_table()

    def _ensure_schedule_table(self) -> None:
        if not self.db or not self.db.bind:
            return
        try:
            ReportSchedule.__table__.create(self.db.bind, checkfirst=True)
        except Exception as exc:  # pragma: no cover - defensive
            logger.warning("Could not ensure report_schedules table: %s", exc)

    def _get_branch_name(self, branch_id: Optional[int]) -> str:
        """Resolve branch_id to branch name."""
        if branch_id is None:
            return "Todas las sucursales"
        try:
            from models.branch import Branch
            branch = self.db.query(Branch).filter(Branch.id == branch_id).first()
            return branch.name if branch else f"Sucursal #{branch_id}"
        except Exception:
            return f"Sucursal #{branch_id}"

    def _get_product_name(self, product_id: Optional[int]) -> str:
        """Resolve product_id to product name."""
        if product_id is None:
            return "Todos los productos"
        try:
            from models.product import Product
            product = self.db.query(Product).filter(Product.id == product_id).first()
            return product.name if product else f"Producto #{product_id}"
        except Exception:
            return f"Producto #{product_id}"

    def _get_user_name(self, user_id: Optional[int]) -> str:
        """Resolve user_id to user name."""
        if user_id is None:
            return "Todos los usuarios"
        try:
            from models.user import User
            user = self.db.query(User).filter(User.id == user_id).first()
            return user.name if user else f"Usuario #{user_id}"
        except Exception:
            return f"Usuario #{user_id}"

    # ─────────────────────────────────────────────────────────────────────────
    # EXP 1-3 · REPORTES BASE (con filtros extendidos)
    # ─────────────────────────────────────────────────────────────────────────

    def generate_inventory_report(
        self,
        branch_id: Optional[int] = None,
        product_id: Optional[int] = None,       # Exp 2
        date_from: Optional[datetime] = None,    # Exp 1 (para last_count_date)
        date_to: Optional[datetime] = None,      # Exp 1
    ) -> Dict[str, Any]:
        """Generate inventory status report with enhanced details."""
        from models.inventory import Inventory
        from models.product import Product
        from models.branch import Branch
        from models.category import Category

        query = (
            self.db.query(Inventory)
            .join(Product)
            .join(Branch)
            .outerjoin(Category, Product.category_id == Category.id)
            .filter(
                Inventory.is_active == True,
                Product.is_active == True,
                Branch.is_active == True,
            )
        )
        if branch_id:
            query = query.filter(Inventory.branch_id == branch_id)
        if product_id:
            query = query.filter(Inventory.product_id == product_id)
        if date_from:
            query = query.filter(Inventory.last_count_date >= date_from)
        if date_to:
            query = query.filter(Inventory.last_count_date <= date_to)

        items = query.all()

        report: Dict[str, Any] = {
            "generated_at": _fmt_date(datetime.utcnow()),
            "filters": {
                "branch_id": branch_id,
                "branch_name": self._get_branch_name(branch_id),
                "product_id": product_id,
                "product_name": self._get_product_name(product_id),
                "date_from": _fmt_date(date_from),
                "date_to": _fmt_date(date_to),
            },
            "total_items": len(items),
            "total_physical_stock": 0,
            "total_digital_stock": 0,
            "total_value": 0.0,
            "discrepancies": [],
            "low_stock": [],
            "by_branch": {},
            "by_category": {},
        }

        for item in items:
            report["total_physical_stock"] += item.physical_stock
            report["total_digital_stock"] += item.digital_stock

            unit_price = item.unit_cost or item.product.unit_price or 0
            item_value = item.digital_stock * unit_price
            report["total_value"] += item_value

            branch_name = item.branch.name
            if branch_name not in report["by_branch"]:
                report["by_branch"][branch_name] = {
                    "items": 0,
                    "physical_stock": 0,
                    "digital_stock": 0,
                    "value": 0.0,
                }
            report["by_branch"][branch_name]["items"] += 1
            report["by_branch"][branch_name]["physical_stock"] += item.physical_stock
            report["by_branch"][branch_name]["digital_stock"] += item.digital_stock
            report["by_branch"][branch_name]["value"] += item_value

            category_name = item.product.category.name if item.product and item.product.category else "Sin categoría"
            if category_name not in report["by_category"]:
                report["by_category"][category_name] = {
                    "items": 0,
                    "digital_stock": 0,
                    "value": 0.0,
                }
            report["by_category"][category_name]["items"] += 1
            report["by_category"][category_name]["digital_stock"] += item.digital_stock
            report["by_category"][category_name]["value"] += item_value

            if item.has_discrepancy:
                report["discrepancies"].append({
                    "product": item.product.name,
                    "sku": item.product.sku if item.product else "—",
                    "branch": branch_name,
                    "physical": item.physical_stock,
                    "digital": item.digital_stock,
                    "difference": item.difference,
                    "percentage": round((abs(item.difference) / max(item.digital_stock, 1)) * 100, 2),
                    "unit_price": unit_price,
                    "impact_value": round(abs(item.difference) * unit_price, 2),
                })

            if item.is_low_stock:
                report["low_stock"].append({
                    "product": item.product.name,
                    "sku": item.product.sku if item.product else "—",
                    "branch": branch_name,
                    "current": item.digital_stock,
                    "min": item.min_stock,
                    "unit_price": unit_price,
                })

        return report


    def generate_movement_report(
        self,
        branch_id: Optional[int] = None,
        product_id: Optional[int] = None,       # Exp 2
        user_id: Optional[int] = None,           # Exp 3
        date_from: Optional[datetime] = None,    # Exp 1
        date_to: Optional[datetime] = None,      # Exp 1
    ) -> Dict[str, Any]:
        """Generate movement history report with enhanced details."""
        from models.movement import Movement
        from models.product import Product
        from models.branch import Branch
        from models.user import User

        query = (
            self.db.query(Movement)
            .join(Product, Movement.product_id == Product.id)
            .join(Branch, Movement.branch_id == Branch.id)
            .join(User, Movement.user_id == User.id)
        )

        if branch_id:
            query = query.filter(
                or_(
                    Movement.branch_id == branch_id,
                    Movement.destination_branch_id == branch_id,
                )
            )
        if product_id:
            query = query.filter(Movement.product_id == product_id)
        if user_id:
            query = query.filter(Movement.user_id == user_id)
        if date_from:
            query = query.filter(Movement.created_at >= date_from)
        if date_to:
            query = query.filter(Movement.created_at <= date_to)

        movements = query.order_by(Movement.created_at.desc()).all()

        report: Dict[str, Any] = {
            "generated_at": _fmt_date(datetime.utcnow()),
            "filters": {
                "branch_id": branch_id,
                "branch_name": self._get_branch_name(branch_id),
                "product_id": product_id,
                "product_name": self._get_product_name(product_id),
                "user_id": user_id,
                "user_name": self._get_user_name(user_id),
                "date_from": _fmt_date(date_from),
                "date_to": _fmt_date(date_to),
            },
            "total_movements": len(movements),
            "total_quantity": 0,
            "total_cost": 0.0,
            "by_type": {},
            "by_state": {},
            "by_branch": {},
            "by_day_of_week": {},
            "movements": [],
        }

        for m in movements:
            mtype = m.movement_type
            report["by_type"].setdefault(mtype, {"count": 0, "total_quantity": 0, "total_cost": 0.0})
            report["by_type"][mtype]["count"] += 1
            report["by_type"][mtype]["total_quantity"] += m.quantity
            report["total_quantity"] += m.quantity
            
            if m.total_cost is not None:
                report["by_type"][mtype]["total_cost"] += m.total_cost
                report["total_cost"] += m.total_cost

            state = m.state
            report["by_state"][state] = report["by_state"].get(state, 0) + 1

            branch_name = m.branch.name if m.branch else "Desconocido"
            b_key = branch_name
            report["by_branch"].setdefault(b_key, {"movements": 0, "quantity": 0, "total_cost": 0.0})
            report["by_branch"][b_key]["movements"] += 1
            report["by_branch"][b_key]["quantity"] += m.quantity
            if m.total_cost is not None:
                report["by_branch"][b_key]["total_cost"] += m.total_cost

            if m.created_at:
                day_name = m.created_at.strftime("%A")
                report["by_day_of_week"][day_name] = report["by_day_of_week"].get(day_name, 0) + 1

            destination_branch_name = "—"
            if m.destination_branch_id:
                dest_branch = self.db.query(Branch).filter(Branch.id == m.destination_branch_id).first()
                destination_branch_name = dest_branch.name if dest_branch else "Desconocido"

            report["movements"].append({
                "id": m.id,
                "date": _fmt_date_readable(m.created_at),
                "product": m.product.name if m.product else "Desconocido",
                "sku": m.product.sku if m.product else "—",
                "quantity": m.quantity,
                "type": m.movement_type,
                "state": m.state,
                "origin_branch": branch_name,
                "destination_branch": destination_branch_name,
                "user": m.user.name if m.user else "Desconocido",
                "unit_cost": m.unit_cost,
                "total_cost": m.total_cost,
            })

        return report


    def generate_discrepancy_report(
        self,
        branch_id: Optional[int] = None,
        product_id: Optional[int] = None,       # Exp 2
        date_from: Optional[datetime] = None,    # Exp 1 (last_count_date)
        date_to: Optional[datetime] = None,      # Exp 1
    ) -> Dict[str, Any]:
        """Generate discrepancy analysis report."""
        from models.inventory import Inventory
        from models.product import Product
        from models.branch import Branch

        query = (
            self.db.query(Inventory)
            .join(Product)
            .join(Branch)
            .filter(
                Inventory.physical_stock != Inventory.digital_stock,
                Inventory.is_active == True,
                Product.is_active == True,
                Branch.is_active == True,
            )
        )
        if branch_id:
            query = query.filter(Inventory.branch_id == branch_id)
        if product_id:
            query = query.filter(Inventory.product_id == product_id)
        if date_from:
            query = query.filter(Inventory.last_count_date >= date_from)
        if date_to:
            query = query.filter(Inventory.last_count_date <= date_to)

        items = query.all()

        report: Dict[str, Any] = {
            "generated_at": _fmt_date(datetime.utcnow()),
            "filters": {
                "branch_id": branch_id,
                "product_id": product_id,
                "date_from": _fmt_date(date_from),
                "date_to": _fmt_date(date_to),
            },
            "total_discrepancies": len(items),
            "items": [],
            "summary": {
                "total_difference": 0,
                "positive_differences": 0,
                "negative_differences": 0,
            },
        }

        for item in items:
            diff = item.difference
            report["summary"]["total_difference"] += abs(diff)
            if diff > 0:
                report["summary"]["positive_differences"] += 1
            elif diff < 0:
                report["summary"]["negative_differences"] += 1

            unit_cost = item.unit_cost or (item.product.unit_price if item.product else None)
            impact_value = abs(diff) * (unit_cost or 0)

            report["items"].append({
                "product": item.product.name,
                "sku": item.product.sku,
                "branch": item.branch.name,
                "physical": item.physical_stock,
                "digital": item.digital_stock,
                "difference": diff,
                "percentage": round((abs(diff) / max(item.digital_stock, 1)) * 100, 2),
                "unit_cost": unit_cost,
                "impact_value": round(impact_value, 2),
            })

        report["items"].sort(key=lambda x: x["percentage"], reverse=True)
        return report


    def generate_kpi_report(
        self,
        branch_id: Optional[int] = None,
        date_from: Optional[datetime] = None,   # Exp 1
        date_to: Optional[datetime] = None,     # Exp 1
    ) -> Dict[str, Any]:
        """Generate KPI report."""
        from modules.dashboard.service import DashboardService
        from modules.inventory.repository import InventoryRepository
        from modules.movements.repository import MovementRepository

        dashboard_service = DashboardService(self.db)

        report: Dict[str, Any] = {
            "generated_at": _fmt_date(datetime.utcnow()),
            "filters": {
                "branch_id": branch_id,
                "date_from": _fmt_date(date_from),
                "date_to": _fmt_date(date_to),
            },
            "kpis": {
                "eri": dashboard_service.calculate_kpi_eri(branch_id),
                "eru": dashboard_service.calculate_kpi_eru(branch_id),
            },
            "metrics": {},
        }

        inventory_repo = InventoryRepository(self.db)
        report["metrics"]["total_physical_stock"] = inventory_repo.get_total_physical_stock(branch_id)
        report["metrics"]["total_digital_stock"] = inventory_repo.get_total_digital_stock(branch_id)
        report["metrics"]["discrepancy_count"] = inventory_repo.get_discrepancy_count(branch_id)
        report["metrics"]["low_stock_count"] = inventory_repo.get_low_stock_count(branch_id)

        movement_repo = MovementRepository(self.db)
        report["metrics"]["pending_movements"] = movement_repo.get_pending_count(branch_id)

        lookback = date_from or (datetime.utcnow() - timedelta(days=30))
        stats = movement_repo.get_stats_by_type(branch_id, lookback)
        report["metrics"]["movement_stats_period"] = stats

        return report

    def generate_user_activity_report(
        self,
        user_id: Optional[int] = None,          # Exp 3
        date_from: Optional[datetime] = None,   # Exp 1
        date_to: Optional[datetime] = None,     # Exp 1
    ) -> Dict[str, Any]:
        """Generate user activity report."""
        from modules.history.service import HistoryService

        history_service = HistoryService(self.db)
        result = history_service.list_history(
            limit=1000,
            user_id=user_id,
            date_from=date_from,
            date_to=date_to,
        )

        report: Dict[str, Any] = {
            "generated_at": _fmt_date(datetime.utcnow()),
            "filters": {
                "user_id": user_id,
                "date_from": _fmt_date(date_from),
                "date_to": _fmt_date(date_to),
            },
            "total_activities": result["total"],
            "by_event_type": {},
            "activities": result["entries"],
        }

        for entry in result["entries"]:
            event_type = entry["event_type"]
            report["by_event_type"][event_type] = (
                report["by_event_type"].get(event_type, 0) + 1
            )

        return report


    # ─────────────────────────────────────────────────────────────────────────
    # EXP 4 · EXPORTACIÓN CSV
    # ─────────────────────────────────────────────────────────────────────────

    def export_to_csv(self, report_data: Dict[str, Any], filename: str) -> str:
        """
        Export report_data to a CSV file at *filename*.

        Nested dicts are flattened with dot-notation keys.
        List values are JSON-encoded into a single cell.
        Returns the filename on success.
        """
        flat = _flatten_dict(report_data)

        # If the report contains an 'items' list we write one row per item
        # plus a summary header block; otherwise we write the flat dict.
        items_list: Optional[List] = None
        for key in ("items", "activities", "products", "branches"):
            val = report_data.get(key)
            if isinstance(val, list) and val:
                items_list = val
                break

        with open(filename, "w", newline="", encoding="utf-8") as fh:
            if items_list and isinstance(items_list[0], dict):
                # --- header rows: scalar fields from the top-level report ---
                meta_writer = csv.writer(fh)
                meta_writer.writerow(["# Reporte generado", flat.get("generated_at", "")])
                for k, v in flat.items():
                    if k not in ("items", "activities", "products", "branches"):
                        meta_writer.writerow([k, v])
                meta_writer.writerow([])  # blank separator

                # --- one row per item ---
                item_writer = csv.DictWriter(
                    fh,
                    fieldnames=list(items_list[0].keys()),
                    extrasaction="ignore",
                )
                item_writer.writeheader()
                item_writer.writerows(items_list)
            else:
                # Simple key-value export
                writer = csv.writer(fh)
                writer.writerow(["campo", "valor"])
                for k, v in flat.items():
                    writer.writerow([k, v])

        logger.info("CSV exported: %s", filename)
        return filename

    # ─────────────────────────────────────────────────────────────────────────
    # EXP 5 · EXPORTACIÓN EXCEL
    # ─────────────────────────────────────────────────────────────────────────

    def export_to_excel(self, report_data: Dict[str, Any], filename: str) -> str:
        """
        Export report_data to an Excel (.xlsx) file at *filename*.

        • Hoja "Resumen"  – campos escalares del reporte (clave / valor).
        • Hoja "Datos"    – lista principal (items / activities / products /
                            branches) con una fila por elemento.
        • Headers en negrita, columnas auto-ajustadas al contenido.
        Returns the filename on success.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError(
                "openpyxl no está instalado. Ejecuta: pip install openpyxl==3.1.5"
            )

        wb = Workbook()

        # ── Hoja 1: Resumen ──────────────────────────────────────────────────
        ws_summary = wb.active
        ws_summary.title = "Resumen"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1F4E78")

        ws_summary["A1"] = "Campo"
        ws_summary["B1"] = "Valor"
        ws_summary["A1"].font = header_font
        ws_summary["B1"].font = header_font
        ws_summary["A1"].fill = header_fill
        ws_summary["B1"].fill = header_fill

        flat = _flatten_dict(report_data)
        list_keys = {"items", "activities", "products", "branches"}
        row = 2
        items_list: Optional[List] = None

        for key in ("items", "activities", "products", "branches"):
            val = report_data.get(key)
            if isinstance(val, list) and val:
                items_list = val
                break

        for k, v in flat.items():
            if k.split(".")[0] in list_keys:
                continue
            ws_summary.cell(row=row, column=1, value=k)
            ws_summary.cell(row=row, column=2, value=str(v) if v is not None else "")
            row += 1

        ws_summary.column_dimensions["A"].width = 35
        ws_summary.column_dimensions["B"].width = 50

        # ── Hoja 2: Datos ────────────────────────────────────────────────────
        if items_list and isinstance(items_list[0], dict):
            ws_data = wb.create_sheet("Datos")
            headers = list(items_list[0].keys())

            for col_idx, header in enumerate(headers, start=1):
                cell = ws_data.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="2E75B6")
                cell.alignment = Alignment(horizontal="center")

            for row_idx, item in enumerate(items_list, start=2):
                for col_idx, header in enumerate(headers, start=1):
                    val = item.get(header)
                    ws_data.cell(row=row_idx, column=col_idx, value=val)

            # Auto-width
            for col_idx, header in enumerate(headers, start=1):
                max_len = max(
                    len(str(header)),
                    max(
                        (len(str(row_item.get(header, "") or "")) for row_item in items_list),
                        default=0,
                    ),
                )
                ws_data.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

        wb.save(filename)
        logger.info("Excel exported: %s", filename)
        return filename

    # ─────────────────────────────────────────────────────────────────────────
    # EXP 6 · REPORTES ADICIONALES / OPCIONALES
    # ─────────────────────────────────────────────────────────────────────────

    def generate_count_session_report(
        self,
        branch_id: Optional[int] = None,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
        status: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Generate a report of inventory count sessions and their outcomes."""
        from models.branch import Branch
        from models.inventory_count_item import InventoryCountItem
        from models.inventory_count_session import InventoryCountSession

        try:
            query = self.db.query(InventoryCountSession)
            if branch_id:
                query = query.filter(InventoryCountSession.branch_id == branch_id)
            if date_from:
                query = query.filter(InventoryCountSession.scheduled_date >= date_from)
            if date_to:
                query = query.filter(InventoryCountSession.scheduled_date <= date_to)
            if status:
                query = query.filter(InventoryCountSession.status == status)

            sessions = query.order_by(InventoryCountSession.scheduled_date.desc()).all()

            branch_cache: Dict[int, str] = {}
            rows: List[Dict[str, Any]] = []
            summary = {
                "total_sessions": len(sessions),
                "completed": 0,
                "in_progress": 0,
                "pending": 0,
                "total_items_counted": 0,
                "total_discrepancies_found": 0,
            }
            by_branch: Dict[str, Dict[str, Any]] = {}

            for session in sessions:
                if session.status == "completed":
                    summary["completed"] += 1
                elif session.status == "in_progress":
                    summary["in_progress"] += 1
                elif session.status == "pending":
                    summary["pending"] += 1
                elif session.status == "cancelled":
                    summary["pending"] += 0

                branch_name = "Sin sucursal"
                if session.branch_id:
                    branch_name = branch_cache.get(session.branch_id)
                    if branch_name is None:
                        branch = self.db.query(Branch).filter(Branch.id == session.branch_id).first()
                        branch_name = branch.name if branch else "Sin sucursal"
                        branch_cache[session.branch_id] = branch_name

                items = list(session.count_items or [])
                items_total = len(items)
                discrepancies = sum(1 for item in items if getattr(item, "is_discrepancy", False))

                summary["total_items_counted"] += items_total
                summary["total_discrepancies_found"] += discrepancies

                by_branch.setdefault(branch_name, {
                    "sessions": 0,
                    "items_counted": 0,
                    "discrepancies": 0,
                })
                by_branch[branch_name]["sessions"] += 1
                by_branch[branch_name]["items_counted"] += items_total
                by_branch[branch_name]["discrepancies"] += discrepancies

                rows.append({
                    "session_id": session.id,
                    "branch_name": branch_name,
                    "scheduled_date": _fmt_date(session.scheduled_date),
                    "status": session.status,
                    "completed_at": _fmt_date(session.completed_at),
                    "validator_count": session.validator_count,
                    "items_total": items_total,
                    "items_with_discrepancy": discrepancies,
                })

            return {
                "generated_at": _fmt_date(datetime.now()),
                "filters": {
                    "branch_id": branch_id,
                    "date_from": _fmt_date(date_from),
                    "date_to": _fmt_date(date_to),
                    "status": status,
                },
                "summary": summary,
                "by_branch": by_branch,
                "sessions": rows,
            }
        except Exception:
            return {
                "generated_at": _fmt_date(datetime.now()),
                "filters": {
                    "branch_id": branch_id,
                    "date_from": _fmt_date(date_from),
                    "date_to": _fmt_date(date_to),
                    "status": status,
                },
                "summary": {
                    "total_sessions": 0,
                    "completed": 0,
                    "in_progress": 0,
                    "pending": 0,
                    "total_items_counted": 0,
                    "total_discrepancies_found": 0,
                },
                "by_branch": {},
                "sessions": [],
            }

    def generate_approval_report(
        self,
        branch_id: Optional[int] = None,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
        approval_level: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Generate a report of transfer approvals and pending requests."""
        from models.branch import Branch
        from models.movement import Movement
        from models.product import Product

        try:
            query = self.db.query(Movement).filter(Movement.movement_type == "transferencia")
            if branch_id:
                query = query.filter(or_(Movement.branch_id == branch_id, Movement.destination_branch_id == branch_id))
            if date_from:
                query = query.filter(Movement.created_at >= date_from)
            if date_to:
                query = query.filter(Movement.created_at <= date_to)
            if approval_level:
                query = query.filter(Movement.approval_level == approval_level)

            movements = query.order_by(Movement.created_at.desc()).all()
            now = datetime.now()
            pending_rows: List[Dict[str, Any]] = []
            by_approver: Dict[str, Dict[str, int]] = {}
            completed_hours: List[float] = []

            for movement in movements:
                product_name = "—"
                product = self.db.query(Product).filter(Product.id == movement.product_id).first()
                if product:
                    product_name = product.name

                origin_branch = "—"
                destination_branch = "—"
                origin = self.db.query(Branch).filter(Branch.id == movement.branch_id).first()
                destination = self.db.query(Branch).filter(Branch.id == movement.destination_branch_id).first()
                if origin:
                    origin_branch = origin.name
                if destination:
                    destination_branch = destination.name

                pending = False
                if movement.requires_approval and movement.approval_level != "none":
                    pending = True
                if movement.requires_approval and not movement.admin_approved and movement.approval_level == "admin":
                    pending = True
                if movement.requires_approval and not movement.manager_approved and movement.approval_level == "manager":
                    pending = True

                if pending:
                    pending_rows.append({
                        "movement_id": movement.id,
                        "product_name": product_name,
                        "origin_branch": origin_branch,
                        "destination_branch": destination_branch,
                        "quantity": movement.quantity,
                        "approval_level": movement.approval_level or "admin",
                        "created_at": _fmt_date(movement.created_at),
                        "waiting_hours": int((now - movement.created_at).total_seconds() // 3600) if movement.created_at else 0,
                    })

                if movement.admin_approved and movement.admin_approved_by:
                    by_approver.setdefault(movement.admin_approved_by, {"approved": 0, "rejected": 0})
                    by_approver[movement.admin_approved_by]["approved"] += 1
                    if movement.admin_approved_at and movement.created_at:
                        completed_hours.append((movement.admin_approved_at - movement.created_at).total_seconds() / 3600)
                elif movement.manager_approved and movement.manager_approved_by:
                    by_approver.setdefault(movement.manager_approved_by, {"approved": 0, "rejected": 0})
                    by_approver[movement.manager_approved_by]["approved"] += 1
                    if movement.manager_approved_at and movement.created_at:
                        completed_hours.append((movement.manager_approved_at - movement.created_at).total_seconds() / 3600)
                elif movement.state == "rechazado":
                    approver_name = movement.validated_by or movement.admin_approved_by or movement.manager_approved_by or "system"
                    by_approver.setdefault(str(approver_name), {"approved": 0, "rejected": 0})
                    by_approver[str(approver_name)]["rejected"] += 1

            return {
                "generated_at": _fmt_date(datetime.now()),
                "filters": {
                    "branch_id": branch_id,
                    "date_from": _fmt_date(date_from),
                    "date_to": _fmt_date(date_to),
                    "approval_level": approval_level,
                },
                "summary": {
                    "total_transfer_requests": len(movements),
                    "pending_admin": sum(1 for item in pending_rows if item["approval_level"] == "admin"),
                    "pending_manager": sum(1 for item in pending_rows if item["approval_level"] == "manager"),
                    "approved": sum(1 for movement in movements if movement.admin_approved or movement.manager_approved),
                    "rejected": sum(1 for movement in movements if movement.state == "rechazado"),
                    "average_approval_time_hours": round(sum(completed_hours) / len(completed_hours), 2) if completed_hours else 0.0,
                },
                "pending_approvals": pending_rows,
                "by_approver": by_approver,
            }
        except Exception:
            return {
                "generated_at": _fmt_date(datetime.now()),
                "filters": {
                    "branch_id": branch_id,
                    "date_from": _fmt_date(date_from),
                    "date_to": _fmt_date(date_to),
                    "approval_level": approval_level,
                },
                "summary": {
                    "total_transfer_requests": 0,
                    "pending_admin": 0,
                    "pending_manager": 0,
                    "approved": 0,
                    "rejected": 0,
                    "average_approval_time_hours": 0.0,
                },
                "pending_approvals": [],
                "by_approver": {},
            }

    def generate_alert_report(
        self,
        branch_id: Optional[int] = None,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
        severity: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Generate alert history statistics."""
        from modules.alerts.service import Alert, AlertService

        try:
            query = self.db.query(Alert)
        except Exception:
            return {
                "generated_at": _fmt_date(datetime.now()),
                "filters": {
                    "branch_id": branch_id,
                    "date_from": _fmt_date(date_from),
                    "date_to": _fmt_date(date_to),
                    "severity": severity,
                },
                "summary": {
                    "total_alerts": 0,
                    "by_severity": {},
                    "by_type": {},
                    "open_alerts": 0,
                    "resolved_alerts": 0,
                    "average_resolution_hours": 0.0,
                },
                "by_branch": {},
                "top_open_alerts": [],
            }
        if branch_id:
            query = query.filter(Alert.branch_id == branch_id)
        if date_from:
            query = query.filter(Alert.created_at >= date_from)
        if date_to:
            query = query.filter(Alert.created_at <= date_to)
        if severity:
            query = query.filter(Alert.severity == severity)

        alerts = query.order_by(Alert.created_at.desc()).all()
        alert_service = AlertService(self.db)
        enriched = [alert_service.enrich_alert(alert.to_dict()) for alert in alerts]

        by_severity: Dict[str, int] = {}
        by_type: Dict[str, int] = {}
        by_branch: Dict[str, Dict[str, Any]] = {}
        resolution_hours: List[float] = []
        top_open: List[Dict[str, Any]] = []
        now = datetime.utcnow()

        for alert in enriched:
            severity_key = alert.get("severity") or "info"
            type_key = alert.get("alert_type") or "unknown"
            by_severity[severity_key] = by_severity.get(severity_key, 0) + 1
            by_type[type_key] = by_type.get(type_key, 0) + 1

            branch_name = alert.get("branch_name") or "Sin sucursal"
            branch_entry = by_branch.setdefault(branch_name, {"total": 0, "open": 0, "critical_open": 0})
            branch_entry["total"] += 1
            if not alert.get("is_resolved"):
                branch_entry["open"] += 1
                if severity_key == "critical":
                    branch_entry["critical_open"] += 1

            if alert.get("is_resolved") and alert.get("resolved_at") and alert.get("created_at"):
                try:
                    resolved = datetime.fromisoformat(alert["resolved_at"].replace("Z", "+00:00"))
                    created = datetime.fromisoformat(alert["created_at"].replace("Z", "+00:00"))
                    resolution_hours.append((resolved - created).total_seconds() / 3600)
                except Exception:
                    pass

            if not alert.get("is_resolved"):
                created_at = alert.get("created_at")
                try:
                    created_dt = datetime.fromisoformat(created_at.replace("Z", "+00:00")) if created_at else now
                    hours_open = round((now - created_dt.replace(tzinfo=None) if created_dt.tzinfo is None else now).total_seconds() / 3600)
                except Exception:
                    hours_open = 0
                top_open.append({
                    "alert_id": alert.get("id"),
                    "type": type_key,
                    "severity": severity_key,
                    "product_name": alert.get("product_name") or "—",
                    "branch_name": branch_name,
                    "created_at": created_at,
                    "hours_open": hours_open,
                })

        top_open.sort(key=lambda item: item["hours_open"], reverse=True)

        return {
            "generated_at": _fmt_date(datetime.now()),
            "filters": {
                "branch_id": branch_id,
                "date_from": _fmt_date(date_from),
                "date_to": _fmt_date(date_to),
                "severity": severity,
            },
            "summary": {
                "total_alerts": len(enriched),
                "by_severity": dict(sorted(by_severity.items())),
                "by_type": dict(sorted(by_type.items())),
                "open_alerts": sum(1 for alert in enriched if not alert.get("is_resolved")),
                "resolved_alerts": sum(1 for alert in enriched if alert.get("is_resolved")),
                "average_resolution_hours": round(sum(resolution_hours) / len(resolution_hours), 2) if resolution_hours else 0.0,
            },
            "by_branch": by_branch,
            "top_open_alerts": top_open[:10],
        }

    def generate_branch_capacity_report(self) -> Dict[str, Any]:
        """Compare branch capacity usage and suggest redistribution opportunities."""
        from models.branch import Branch
        from models.inventory import Inventory

        branches = self.db.query(Branch).filter(Branch.is_active == True).order_by(Branch.name).all()
        rows: List[Dict[str, Any]] = []

        for branch in branches:
            inventory_items = self.db.query(Inventory).filter(Inventory.branch_id == branch.id, Inventory.is_active == True).all()
            current_skus = len({item.product_id for item in inventory_items if item.product_id})
            max_products = branch.max_products
            usage_percent = round((current_skus / max_products) * 100, 2) if max_products else None
            if max_products is None:
                status = "no_limit"
            elif usage_percent is None:
                status = "no_limit"
            elif usage_percent >= 90:
                status = "critical"
            elif usage_percent >= 70:
                status = "warning"
            else:
                status = "ok"

            last_update = None
            for item in inventory_items:
                candidate = item.last_count_date or item.updated_at
                if candidate and (last_update is None or candidate > last_update):
                    last_update = candidate

            rows.append({
                "branch_id": branch.id,
                "branch_name": branch.name,
                "current_skus": current_skus,
                "max_products": max_products,
                "usage_percent": usage_percent,
                "status": status,
                "last_inventory_update": _fmt_date(last_update),
            })

        suggestions: List[Dict[str, Any]] = []
        critical_branches = [row for row in rows if row["status"] == "critical"]
        low_usage_branches = [row for row in rows if row["status"] in ("ok", "warning") and row["usage_percent"] is not None and row["usage_percent"] < 60]
        for source in critical_branches[:3]:
            for target in low_usage_branches[:3]:
                if source["branch_name"] != target["branch_name"]:
                    suggestions.append({
                        "from_branch": source["branch_name"],
                        "to_branch": target["branch_name"],
                        "reason": f"{source['branch_name']} al {source['usage_percent']}%, {target['branch_name']} al {target['usage_percent']}%",
                    })

        return {
            "generated_at": _fmt_date(datetime.now()),
            "summary": {
                "total_branches": len(rows),
                "with_capacity_limit": sum(1 for row in rows if row["max_products"] is not None),
                "without_capacity_limit": sum(1 for row in rows if row["max_products"] is None),
                "avg_usage_percent": round(sum(row["usage_percent"] or 0 for row in rows) / max(1, len(rows)), 2),
            },
            "branches": rows,
            "transfer_suggestions": suggestions,
        }

    def generate_config_change_report(
        self,
        branch_id: Optional[int] = None,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
        entity_type: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Generate a report of configuration and product changes."""
        from models.branch import Branch
        from models.branch_config_history import BranchConfigHistory
        from models.product import Product
        from models.product_change_history import ProductChangeHistory
        from modules.history.service import HistoryService

        changes: List[Dict[str, Any]] = []

        try:
            self.db.query(BranchConfigHistory).count()
        except Exception:
            return {
                "generated_at": _fmt_date(datetime.now()),
                "filters": {
                    "branch_id": branch_id,
                    "date_from": _fmt_date(date_from),
                    "date_to": _fmt_date(date_to),
                    "entity_type": entity_type,
                },
                "summary": {"total_changes": 0, "by_entity_type": {}},
                "changes": [],
            }

        if entity_type in (None, "branch_config"):
            query = self.db.query(BranchConfigHistory)
            if branch_id:
                query = query.filter(BranchConfigHistory.branch_id == branch_id)
            if date_from:
                query = query.filter(BranchConfigHistory.changed_at >= date_from)
            if date_to:
                query = query.filter(BranchConfigHistory.changed_at <= date_to)
            for record in query.order_by(BranchConfigHistory.changed_at.desc()).all():
                branch_name = "—"
                if record.branch_id:
                    branch = self.db.query(Branch).filter(Branch.id == record.branch_id).first()
                    branch_name = branch.name if branch else str(record.branch_id)
                changes.append({
                    "id": record.id,
                    "entity_type": "branch_config",
                    "entity_name": branch_name,
                    "field_name": record.field_name,
                    "old_value": record.old_value,
                    "new_value": record.new_value,
                    "changed_by": record.changed_by,
                    "changed_at": _fmt_date(record.changed_at),
                })

        if entity_type in (None, "product"):
            query = self.db.query(ProductChangeHistory)
            if date_from:
                query = query.filter(ProductChangeHistory.changed_at >= date_from)
            if date_to:
                query = query.filter(ProductChangeHistory.changed_at <= date_to)
            for record in query.order_by(ProductChangeHistory.changed_at.desc()).all():
                product_name = "—"
                if record.product_id:
                    product = self.db.query(Product).filter(Product.id == record.product_id).first()
                    product_name = product.name if product else str(record.product_id)
                changes.append({
                    "id": record.id,
                    "entity_type": "product",
                    "entity_name": product_name,
                    "field_name": record.field_name,
                    "old_value": record.old_value,
                    "new_value": record.new_value,
                    "changed_by": record.changed_by_name,
                    "changed_at": _fmt_date(record.changed_at),
                })

        if entity_type in (None, "branch"):
            history_svc = HistoryService(self.db)
            result = history_svc.list_history(
                limit=500,
                entity_type="branch",
                branch_id=branch_id,
                date_from=date_from,
                date_to=date_to,
                enrich=True,
            )
            for entry in result.get("entries", []):
                changes.append({
                    "id": entry.get("id"),
                    "entity_type": "branch",
                    "entity_name": entry.get("entity_name") or "—",
                    "field_name": entry.get("event_type") or "—",
                    "old_value": None,
                    "new_value": None,
                    "changed_by": entry.get("user_name"),
                    "changed_at": entry.get("created_at"),
                })

        changes.sort(key=lambda item: item.get("changed_at") or "", reverse=True)
        by_entity_type: Dict[str, int] = {}
        for change in changes:
            by_entity_type[change["entity_type"]] = by_entity_type.get(change["entity_type"], 0) + 1

        return {
            "generated_at": _fmt_date(datetime.now()),
            "filters": {
                "branch_id": branch_id,
                "date_from": _fmt_date(date_from),
                "date_to": _fmt_date(date_to),
                "entity_type": entity_type,
            },
            "summary": {
                "total_changes": len(changes),
                "by_entity_type": by_entity_type,
            },
            "changes": changes,
        }

    def generate_batch_report(
        self,
        branch_id: Optional[int] = None,
        days_until_expiry: Optional[int] = None,
    ) -> Dict[str, Any]:
        """Generate an inventory batch report for expiring and expired lots."""
        from models.branch import Branch
        from models.inventory import Inventory
        from models.inventory_batch import InventoryBatch
        from models.product import Product

        query = (
            self.db.query(InventoryBatch)
            .join(Inventory)
            .join(Product)
            .join(Branch)
            .filter(
                Inventory.is_active == True,
                Product.is_active == True,
                Branch.is_active == True,
            )
        )
        if branch_id:
            query = query.filter(Inventory.branch_id == branch_id)

        batches = query.order_by(InventoryBatch.expiration_date.asc().nulls_last()).all()
        today = datetime.now().date()
        expiring_batches: List[Dict[str, Any]] = []
        expired_batches: List[Dict[str, Any]] = []
        by_branch: Dict[str, Dict[str, int]] = {}
        total_quantity_at_risk = 0

        for batch in batches:
            inventory = self.db.query(Inventory).filter(Inventory.id == batch.inventory_id).first()
            branch_name = "Sin sucursal"
            if inventory and inventory.branch_id:
                branch = self.db.query(Branch).filter(Branch.id == inventory.branch_id).first()
                branch_name = branch.name if branch else "Sin sucursal"
            if inventory and inventory.product_id:
                product = self.db.query(Product).filter(Product.id == inventory.product_id).first()
                product_name = product.name if product else "—"
            else:
                product_name = "—"

            if not batch.expiration_date:
                continue

            days = (batch.expiration_date - today).days
            if days < 0:
                expired_batches.append({
                    "batch_id": batch.id,
                    "batch_number": batch.batch_number,
                    "product_name": product_name,
                    "branch_name": branch_name,
                    "expiration_date": batch.expiration_date.isoformat(),
                    "days_until_expiry": days,
                    "quantity": batch.quantity,
                })
                total_quantity_at_risk += batch.quantity
                by_branch.setdefault(branch_name, {"total": 0, "expiring": 0, "expired": 0})
                by_branch[branch_name]["expired"] += 1
            elif days_until_expiry is None or days <= days_until_expiry:
                expiring_batches.append({
                    "batch_id": batch.id,
                    "batch_number": batch.batch_number,
                    "product_name": product_name,
                    "branch_name": branch_name,
                    "expiration_date": batch.expiration_date.isoformat(),
                    "days_until_expiry": days,
                    "quantity": batch.quantity,
                })
                total_quantity_at_risk += batch.quantity
                by_branch.setdefault(branch_name, {"total": 0, "expiring": 0, "expired": 0})
                by_branch[branch_name]["expiring"] += 1

            by_branch.setdefault(branch_name, {"total": 0, "expiring": 0, "expired": 0})
            by_branch[branch_name]["total"] += 1

        return {
            "generated_at": _fmt_date(datetime.now()),
            "filters": {"branch_id": branch_id, "days_until_expiry": days_until_expiry},
            "summary": {
                "total_batches": len(batches),
                "expiring_soon": len(expiring_batches),
                "expired": len(expired_batches),
                "total_quantity_at_risk": total_quantity_at_risk,
            },
            "expiring_batches": expiring_batches,
            "expired_batches": expired_batches,
            "by_branch": by_branch,
        }

    def generate_kit_report(self, branch_id: Optional[int] = None) -> Dict[str, Any]:
        """Generate a report of products that are kits and their components."""
        from models.inventory import Inventory
        from models.product import Product

        query = self.db.query(Product).filter(Product.is_active == True, Product.is_kit == True).order_by(Product.name)
        if branch_id:
            query = query.filter(Product.is_active == True)

        kits: List[Dict[str, Any]] = []
        total_components = 0
        for product in query.all():
            components = []
            for component in product.kit_components:
                total_components += 1
                components.append({
                    "name": component.component_product.name if component.component_product else "—",
                    "sku": component.component_product.sku if component.component_product else "—",
                    "quantity_needed": component.quantity,
                })

            if branch_id:
                inventory_item = self.db.query(Inventory).filter(Inventory.branch_id == branch_id, Inventory.product_id == product.id).first()
                total_stock_kit = inventory_item.digital_stock if inventory_item else 0
            else:
                total_stock_kit = self.db.query(func.sum(Inventory.digital_stock)).filter(Inventory.product_id == product.id, Inventory.is_active == True).scalar() or 0

            if total_stock_kit <= 0:
                stock_status = "out"
            elif total_stock_kit < max(1, len(components)):
                stock_status = "low"
            else:
                stock_status = "ok"

            kits.append({
                "product_name": product.name,
                "sku": product.sku,
                "components": components,
                "total_stock_kit": int(total_stock_kit),
                "stock_status": stock_status,
            })

        return {
            "generated_at": _fmt_date(datetime.now()),
            "summary": {
                "total_kits": len(kits),
                "total_components": total_components,
            },
            "kits": kits,
        }

    def generate_abc_analysis_report(
        self,
        branch_id: Optional[int] = None,
        days_without_movement: int = 90,
    ) -> Dict[str, Any]:
        """Generate ABC product classification and dead stock analysis."""
        from models.branch import Branch
        from models.inventory import Inventory
        from models.movement import Movement
        from models.product import Product

        query = (
            self.db.query(Inventory)
            .join(Product)
            .join(Branch)
            .filter(
                Inventory.is_active == True,
                Product.is_active == True,
                Branch.is_active == True,
            )
        )
        if branch_id:
            query = query.filter(Inventory.branch_id == branch_id)

        inventory_items = query.all()
        product_values: Dict[int, Dict[str, Any]] = {}
        for item in inventory_items:
            value = item.digital_stock * (item.unit_cost if item.unit_cost is not None else item.product.unit_price or 0)
            entry = product_values.setdefault(item.product_id, {
                "product_name": item.product.name,
                "sku": item.product.sku,
                "total_value": 0.0,
                "current_stock": 0,
            })
            entry["total_value"] += value
            entry["current_stock"] += item.digital_stock

        ranked_products = sorted(product_values.values(), key=lambda item: item["total_value"], reverse=True)
        total_value = sum(item["total_value"] for item in ranked_products)
        class_a_products: List[Dict[str, Any]] = []
        class_b_products: List[Dict[str, Any]] = []
        class_c_products: List[Dict[str, Any]] = []
        cumulative_value = 0.0

        for item in ranked_products:
            percent_of_total = round((item["total_value"] / total_value * 100) if total_value else 0.0, 2)
            cumulative_value += item["total_value"]
            if cumulative_value <= total_value * 0.8:
                class_name = "A"
                class_a_products.append({"product_name": item["product_name"], "sku": item["sku"], "total_value": round(item["total_value"], 2), "percent_of_total": percent_of_total})
            elif cumulative_value <= total_value * 0.95:
                class_name = "B"
                class_b_products.append({"product_name": item["product_name"], "sku": item["sku"], "total_value": round(item["total_value"], 2), "percent_of_total": percent_of_total})
            else:
                class_name = "C"
                class_c_products.append({"product_name": item["product_name"], "sku": item["sku"], "total_value": round(item["total_value"], 2), "percent_of_total": percent_of_total})

        cutoff = datetime.now() - timedelta(days=days_without_movement)
        dead_stock: List[Dict[str, Any]] = []
        for product_id, data in product_values.items():
            if data["current_stock"] <= 0:
                continue
            has_movement = self.db.query(Movement).filter(
                Movement.product_id == product_id,
                Movement.created_at >= cutoff,
            ).first() is not None
            if not has_movement:
                dead_stock.append({
                    "product_name": data["product_name"],
                    "sku": data["sku"],
                    "days_without_movement": days_without_movement,
                    "current_stock": data["current_stock"],
                })

        return {
            "generated_at": _fmt_date(datetime.now()),
            "filters": {"days_without_movement": days_without_movement},
            "summary": {
                "class_a_count": len(class_a_products),
                "class_b_count": len(class_b_products),
                "class_c_count": len(class_c_products),
                "dead_stock_count": len(dead_stock),
                "dead_stock_value": round(sum(item["total_value"] for item in ranked_products if item["product_name"] in {d["product_name"] for d in dead_stock}), 2),
            },
            "class_a_products": class_a_products,
            "dead_stock": dead_stock,
        }

    def create_report_schedule(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Create a new recurring report schedule."""
        self._ensure_schedule_table()
        payload = data or {}
        name = (payload.get("name") or "").strip()
        report_type = (payload.get("report_type") or "").strip()
        schedule_type = (payload.get("schedule_type") or "daily").strip()
        schedule_time = (payload.get("schedule_time") or "00:00").strip()
        if not name or not report_type:
            raise ValueError("name y report_type son obligatorios")

        schedule = ReportSchedule(
            name=name,
            report_type=report_type,
            schedule_type=schedule_type,
            schedule_time=schedule_time,
            schedule_day_of_week=payload.get("schedule_day_of_week"),
            schedule_day_of_month=payload.get("schedule_day_of_month"),
            parameters=json.dumps(payload.get("parameters") or {}, ensure_ascii=False),
            created_by_user_id=payload.get("created_by_user_id"),
            is_active=bool(payload.get("is_active", True)),
        )
        schedule.next_run_at = self._calculate_next_run(schedule)
        self.db.add(schedule)
        self.db.commit()
        self.db.refresh(schedule)
        return schedule.to_dict()

    def update_report_schedule(self, schedule_id: int, data: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """Update an existing report schedule."""
        self._ensure_schedule_table()
        schedule = self.db.query(ReportSchedule).filter(ReportSchedule.id == schedule_id).first()
        if not schedule:
            return None
        payload = data or {}
        for field in ("name", "report_type", "schedule_type", "schedule_time", "schedule_day_of_week", "schedule_day_of_month", "is_active"):
            if field in payload:
                setattr(schedule, field, payload[field])
        if "parameters" in payload:
            schedule.parameters = json.dumps(payload["parameters"] or {}, ensure_ascii=False)
        schedule.next_run_at = self._calculate_next_run(schedule)
        self.db.commit()
        self.db.refresh(schedule)
        return schedule.to_dict()

    def delete_report_schedule(self, schedule_id: int) -> bool:
        """Delete a report schedule."""
        self._ensure_schedule_table()
        schedule = self.db.query(ReportSchedule).filter(ReportSchedule.id == schedule_id).first()
        if not schedule:
            return False
        self.db.delete(schedule)
        self.db.commit()
        return True

    def list_report_schedules(self, include_inactive: bool = False) -> List[Dict[str, Any]]:
        """List report schedules."""
        self._ensure_schedule_table()
        query = self.db.query(ReportSchedule)
        if not include_inactive:
            query = query.filter(ReportSchedule.is_active == True)
        return [schedule.to_dict() for schedule in query.order_by(ReportSchedule.created_at.desc()).all()]

    def run_report_schedule(self, schedule_id: int) -> Dict[str, Any]:
        """Run a scheduled report now and persist its output to disk."""
        self._ensure_schedule_table()
        schedule = self.db.query(ReportSchedule).filter(ReportSchedule.id == schedule_id).first()
        if not schedule:
            raise ValueError(f"Schedule {schedule_id} not found")

        report_type = schedule.report_type
        params = json.loads(schedule.parameters or "{}") if schedule.parameters else {}
        report_data = self._generate_report_data(report_type, params)
        file_name = self._generate_scheduled_report_name(schedule, datetime.now())
        output_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "generated_reports")
        os.makedirs(output_dir, exist_ok=True)
        file_path = os.path.join(output_dir, f"{file_name}.txt")
        with open(file_path, "w", encoding="utf-8") as fh:
            fh.write(self.format_report_for_display(report_type, report_data))

        schedule.last_run_at = datetime.now()
        schedule.next_run_at = self._calculate_next_run(schedule)
        self.db.commit()
        self.db.refresh(schedule)
        return {"schedule_id": schedule.id, "report_type": report_type, "file_path": file_path, "name": file_name}

    def process_due_schedules(self) -> int:
        """Run every active schedule whose next execution time has arrived."""
        self._ensure_schedule_table()
        now = datetime.utcnow()
        schedules = self.db.query(ReportSchedule).filter(ReportSchedule.is_active == True, ReportSchedule.next_run_at <= now).all()
        for schedule in schedules:
            try:
                self.run_report_schedule(schedule.id)
            except Exception as exc:  # pragma: no cover - defensive
                logger.warning("Scheduled report failed for %s: %s", schedule.id, exc)
        return len(schedules)

    def _calculate_next_run(self, schedule: Any) -> Optional[datetime]:
        if isinstance(schedule, dict):
            schedule_type = (schedule.get("schedule_type") or "daily").lower()
            schedule_time = schedule.get("schedule_time") or "00:00"
            day_of_week = schedule.get("schedule_day_of_week")
            day_of_month = schedule.get("schedule_day_of_month")
        else:
            schedule_type = (getattr(schedule, "schedule_type", "daily") or "daily").lower()
            schedule_time = getattr(schedule, "schedule_time", "00:00") or "00:00"
            day_of_week = getattr(schedule, "schedule_day_of_week", None)
            day_of_month = getattr(schedule, "schedule_day_of_month", None)

        hour, minute = map(int, str(schedule_time).split(":", 1))
        base = datetime.utcnow().replace(second=0, microsecond=0)
        candidate = base.replace(hour=hour, minute=minute)
        if candidate <= base:
            candidate = candidate + timedelta(days=1)

        if schedule_type == "weekly":
            target_weekday = int(day_of_week or base.weekday())
            while candidate.weekday() != target_weekday:
                candidate = candidate + timedelta(days=1)
        elif schedule_type == "monthly":
            day = int(day_of_month or base.day)
            while candidate.day != day:
                candidate = candidate + timedelta(days=1)
                if candidate.month != base.month and candidate.day != day:
                    break
            if candidate.day != day:
                candidate = candidate.replace(day=1) + timedelta(days=32)
                candidate = candidate.replace(day=day)
        return candidate

    def _generate_scheduled_report_name(self, schedule: Any, generated_at: datetime) -> str:
        base_name = re.sub(r"[^A-Za-z0-9]+", "_", getattr(schedule, "name", "reporte") or "reporte").strip("_").lower() or "reporte"
        return f"{base_name}_{generated_at.strftime('%Y-%m-%d_%H-%M')}"

    def _generate_report_data(self, report_type: str, params: Dict[str, Any]) -> Dict[str, Any]:
        branch_id = params.get("branch_id")
        date_from = params.get("date_from")
        date_to = params.get("date_to")
        if isinstance(date_from, str):
            date_from = datetime.fromisoformat(date_from)
        if isinstance(date_to, str):
            date_to = datetime.fromisoformat(date_to)

        if report_type == "count_sessions":
            return self.generate_count_session_report(branch_id=branch_id, date_from=date_from, date_to=date_to, status=params.get("status"))
        if report_type == "approvals":
            return self.generate_approval_report(branch_id=branch_id, date_from=date_from, date_to=date_to, approval_level=params.get("approval_level"))
        if report_type == "alerts":
            return self.generate_alert_report(branch_id=branch_id, date_from=date_from, date_to=date_to, severity=params.get("severity"))
        if report_type == "branch_capacity":
            return self.generate_branch_capacity_report()
        if report_type == "config_changes":
            return self.generate_config_change_report(branch_id=branch_id, date_from=date_from, date_to=date_to, entity_type=params.get("entity_type"))
        if report_type == "batches":
            return self.generate_batch_report(branch_id=branch_id, days_until_expiry=params.get("days_until_expiry"))
        if report_type == "kits":
            return self.generate_kit_report(branch_id=branch_id)
        if report_type == "abc_analysis":
            return self.generate_abc_analysis_report(branch_id=branch_id, days_without_movement=params.get("days_without_movement", 90))
        raise ValueError(f"Tipo de reporte no soportado para programación: {report_type}")

    def format_report_for_display(
        self, report_type: str, report_data: Dict[str, Any]
    ) -> str:
        """
        Convert a report dict into a human-readable text block.

        Returns a plain-text string with borders, aligned columns and totals.
        """
        lines: List[str] = []

        def h1(text: str) -> None:
            bar = "═" * (len(text) + 4)
            lines.append(f"╔{bar}╗")
            lines.append(f"║  {text}  ║")
            lines.append(f"╚{bar}╝")

        def h2(text: str) -> None:
            lines.append(f"\n  ┌─ {text} {'─' * max(0, 60 - len(text))}┐")

        def kv(label: str, value: Any, indent: int = 4) -> None:
            lines.append(f"{' ' * indent}{label:<30} {value}")

        def table(headers: List[str], rows: List[List[Any]], indent: int = 4) -> None:
            col_widths = [
                max(len(str(h)), max((len(str(r[i] or "")) for r in rows), default=0))
                for i, h in enumerate(headers)
            ]
            sep = " " * indent + "+" + "+".join("-" * (w + 2) for w in col_widths) + "+"
            header_row = (
                " " * indent
                + "| "
                + " | ".join(str(h).ljust(w) for h, w in zip(headers, col_widths))
                + " |"
            )
            lines.append(sep)
            lines.append(header_row)
            lines.append(sep)
            for row in rows:
                lines.append(
                    " " * indent
                    + "| "
                    + " | ".join(str(c or "").ljust(w) for c, w in zip(row, col_widths))
                    + " |"
                )
            lines.append(sep)

        def alert(text: str, indent: int = 4) -> None:
            lines.append(f"{' ' * indent}⚠ {text}")

        def insight(text: str, indent: int = 4) -> None:
            lines.append(f"{' ' * indent}💡 {text}")

        def success(text: str, indent: int = 4) -> None:
            lines.append(f"{' ' * indent}✅ {text}")

        def warning(text: str, indent: int = 4) -> None:
            lines.append(f"{' ' * indent}🟡 {text}")

        def critical(text: str, indent: int = 4) -> None:
            lines.append(f"{' ' * indent}🔴 {text}")

        def recommendation(text: str, indent: int = 4) -> None:
            lines.append(f"{' ' * indent}▶ {text}")

        def metric_card(label: str, value: str, status: str = "neutral", indent: int = 4) -> None:
            """Display a metric with visual status indicator."""
            icons = {"good": "🟢", "warning": "🟡", "critical": "🔴", "neutral": "⚪"}
            icon = icons.get(status, "⚪")
            lines.append(f"{' ' * indent}{icon} {label:<28} {value}")

        def executive_summary(text: str) -> None:
            lines.append(f"\n  ┌─ RESUMEN EJECUTIVO {'─' * 50}┐")
            for line in text.strip().split('\n'):
                lines.append(f"  │ {line}")
            lines.append(f"  └{'─' * 70}┘")

        TITLES = {
            "inventory": "Reporte de Inventario",
            "movements": "Reporte de Movimientos",
            "discrepancies": "Reporte de Discrepancias",
            "kpis": "Reporte de KPIs",
            "user_activity": "Actividad por Usuario",
            "top_products": "Top Productos",
            "branch_efficiency": "Eficiencia por Sucursal",
            "inventory_value": "Valor de Inventario",
            "transfers": "Reporte de Transferencias",
            "trends": "Tendencias",
            "comparison": "Comparación de Períodos",
            "count_sessions": "Conteos de Inventario",
            "approvals": "Aprobaciones",
            "alerts": "Historial de Alertas",
            "branch_capacity": "Capacidad de Sucursal",
            "config_changes": "Cambios de Configuración",
            "batches": "Reporte de Lotes",
            "kits": "Reporte de Kits",
            "abc_analysis": "Análisis ABC y Dead Stock",
        }
        h1(TITLES.get(report_type, "Reporte"))
        generated_at = report_data.get("generated_at", "—")
        if generated_at and generated_at != "—":
            try:
                from datetime import datetime
                dt = datetime.fromisoformat(generated_at.replace("Z", "+00:00"))
                kv("Generado en:", _fmt_date_readable(dt), indent=2)
            except Exception:
                kv("Generado en:", generated_at, indent=2)
        else:
            kv("Generado en:", generated_at, indent=2)

        filters = report_data.get("filters", {})
        if any(v for v in filters.values()):
            h2("Filtros aplicados")
            for k, v in filters.items():
                if v is not None and v != "—":
                    label = k.replace("_", " ").replace("id", "ID").replace("name", "Nombre").title()
                    if k.endswith("_name") and k.replace("_name", "_id") in filters:
                        continue
                    if k.endswith("_id") and k.replace("_id", "_name") in filters:
                        continue
                    if k.endswith("_id"):
                        continue
                    kv(label + ":", v)

# ── Inventario ────────────────────────────────────────────────────────
        if report_type == "inventory":
            total_items = report_data.get("total_items", 0)
            total_physical = report_data.get("total_physical_stock", 0)
            total_digital = report_data.get("total_digital_stock", 0)
            total_value = report_data.get("total_value", 0)
            discrepancies = report_data.get("discrepancies", [])
            low_stock = report_data.get("low_stock", [])
            
            discrepancy_count = len(discrepancies)
            low_stock_count = len(low_stock)
            discrepancy_pct = (discrepancy_count / total_items * 100) if total_items > 0 else 0
            low_stock_pct = (low_stock_count / total_items * 100) if total_items > 0 else 0
            avg_value_per_item = (total_value / total_items) if total_items > 0 else 0
            
            # Executive Summary
            exec_lines = []
            if total_items > 0:
                exec_lines.append(f"Inventario total: {total_items} artículos únicos, {total_digital:,} unidades en stock digital.")
                if total_value > 0:
                    exec_lines.append(f"Valor total de inventario: ${total_value:,.2f} (promedio ${avg_value_per_item:.2f}/artículo).")
                
                # Discrepancy assessment
                if discrepancy_count == 0:
                    exec_lines.append("✅ Integridad de inventario: SIN discrepancias detectadas.")
                elif discrepancy_pct < 5:
                    exec_lines.append(f"⚠️  {discrepancy_count} artículos con discrepancia ({discrepancy_pct:.1f}%) - Tolerable, monitorear.")
                else:
                    exec_lines.append(f"🔴 {discrepancy_count} artículos con discrepancia ({discrepancy_pct:.1f}%) - Requiere investigación inmediata.")
                
                # Low stock assessment
                if low_stock_count == 0:
                    exec_lines.append("✅ Niveles de stock: TODOS los productos por encima de mínimos.")
                elif low_stock_pct < 10:
                    exec_lines.append(f"⚠️  {low_stock_count} productos bajo mínimo ({low_stock_pct:.1f}%) - Planear reposición.")
                else:
                    exec_lines.append(f"🔴 {low_stock_count} productos bajo mínimo ({low_stock_pct:.1f}%) - REPOSICIÓN URGENTE.")
            else:
                exec_lines.append("No hay artículos en inventario para el período/filtros seleccionados.")
            
            executive_summary("\n".join(exec_lines))
            
            # Key Metrics Dashboard
            h2("📊 Indicadores Clave")
            metric_card("Artículos únicos", f"{total_items:,}")
            metric_card("Stock físico total", f"{total_physical:,}")
            metric_card("Stock digital total", f"{total_digital:,}")
            if total_value:
                metric_card("Valor total inventario", f"${total_value:,.2f}")
            metric_card("Discrepancias", f"{discrepancy_count} ({discrepancy_pct:.1f}%)", 
                       "good" if discrepancy_pct == 0 else ("warning" if discrepancy_pct < 5 else "critical"))
            metric_card("Stock bajo mínimo", f"{low_stock_count} ({low_stock_pct:.1f}%)",
                       "good" if low_stock_pct == 0 else ("warning" if low_stock_pct < 10 else "critical"))
            
            # Recommendations
            recommendations = []
            if discrepancy_count > 0:
                top_disc = max(discrepancies, key=lambda x: x.get('impact_value', 0)) if discrepancies else None
                if top_disc:
                    recommendations.append(f"Investigar discrepancia mayor: {top_disc['product']} ({top_disc['branch']}) - Impacto: ${top_disc.get('impact_value', 0):,.2f}")
            if low_stock_count > 0:
                recommendations.append(f"Generar órdenes de compra para {low_stock_count} productos bajo mínimo")
            if discrepancy_pct > 5:
                recommendations.append("Programar conteo físico completo en sucursales con mayor discrepancia")
            if total_value > 0 and avg_value_per_item > 1000:
                recommendations.append("Revisar cobertura de seguro para inventario de alto valor")
            
            if recommendations:
                h2("🎯 Recomendaciones")
                for i, rec in enumerate(recommendations, 1):
                    recommendation(f"{i}. {rec}")
            
            h2("Resumen")
            kv("Total artículos:", total_items)
            kv("Stock físico total:", total_physical)
            kv("Stock digital total:", total_digital)
            if total_value:
                kv("Valor total inventario:", f"${total_value:,.2f}")
            
            if report_data.get("by_branch"):
                h2("Por sucursal")
                tbl_rows = []
                for b, v in report_data["by_branch"].items():
                    row = [b, v["items"], v["physical_stock"], v["digital_stock"]]
                    if "value" in v:
                        row.append(f"${v['value']:,.2f}")
                    tbl_rows.append(row)
                headers = ["Sucursal", "Artículos", "Físico", "Digital"]
                if "value" in list(report_data["by_branch"].values())[0]:
                    headers.append("Valor")
                table(headers, tbl_rows)
            
            if report_data.get("by_category"):
                h2("Por categoría")
                table(
                    ["Categoría", "Artículos", "Stock", "Valor"],
                    [[c, v["items"], v["digital_stock"], f"${v['value']:,.2f}"]
                     for c, v in report_data["by_category"].items()],
                )
            
            if discrepancies:
                h2(f"⚠ Discrepancias ({len(discrepancies)} ítems)")
                table(
                    ["Producto", "SKU", "Sucursal", "Físico", "Digital", "Diferencia", "%", "Valor impacto"],
                    [[i["product"], i["sku"], i["branch"],
                      i["physical"], i["digital"], i["difference"], f"{i['percentage']}%",
                      f"${i.get('impact_value', 0):,.2f}"]
                      for i in discrepancies[:50]],
                )
            
            if low_stock:
                h2(f"🟡 Stock bajo ({len(low_stock)} artículos)")
                low_stock_rows = []
                for i in low_stock:
                    row = [i["product"], i["sku"], i["branch"], i["current"], i["min"]]
                    if "unit_price" in i:
                        row.append(f"${i['unit_price']:.2f}")
                    low_stock_rows.append(row)
                headers = ["Producto", "SKU", "Sucursal", "Actual", "Mínimo"]
                if low_stock and "unit_price" in low_stock[0]:
                    headers.append("Costo")
                table(headers, low_stock_rows)

# ── Movimientos ───────────────────────────────────────────────────────
        elif report_type == "movements":
            total_movements = report_data.get("total_movements", 0)
            total_quantity = report_data.get("total_quantity", 0)
            by_state = report_data.get("by_state", {})
            by_type = report_data.get("by_type", {})
            
            validated_count = by_state.get("validado", 0)
            pending_count = by_state.get("pendiente", 0)
            rejected_count = by_state.get("rechazado", 0)
            validated_pct = (validated_count / total_movements * 100) if total_movements > 0 else 0
            pending_pct = (pending_count / total_movements * 100) if total_movements > 0 else 0
            rejected_pct = (rejected_count / total_movements * 100) if total_movements > 0 else 0
            avg_per_movement = total_quantity / total_movements if total_movements > 0 else 0
            total_cost = report_data.get("total_cost", 0)
            
            # Executive Summary
            exec_lines = []
            if total_movements > 0:
                exec_lines.append(f"Se registraron {total_movements:,} movimientos con {total_quantity:,} unidades totales.")
                exec_lines.append(f"Promedio: {avg_per_movement:.1f} unidades por movimiento.")
                if total_cost > 0:
                    exec_lines.append(f"Costo total de movimientos: ${total_cost:,.2f}.")
                
                # Validation status
                if validated_pct == 100:
                    exec_lines.append("✅ Validación: TODOS los movimientos están validados.")
                elif validated_pct >= 80:
                    exec_lines.append(f"✅ Validación: {validated_pct:.0f}% validados ({validated_count}/{total_movements}) - Buena cobertura.")
                elif validated_pct >= 50:
                    exec_lines.append(f"⚠️  Validación: {validated_pct:.0f}% validados ({validated_count}/{total_movements}) - Revisar pendientes.")
                else:
                    exec_lines.append(f"🔴 Validación: Solo {validated_pct:.0f}% validados ({validated_count}/{total_movements}) - CRÍTICO.")
                
                # Pending/Rejected assessment
                if pending_count > 0:
                    exec_lines.append(f"⏳ {pending_count} movimientos pendientes ({pending_pct:.1f}%) requieren atención.")
                if rejected_count > 0:
                    exec_lines.append(f"❌ {rejected_count} movimientos rechazados ({rejected_pct:.1f}%) - Analizar causas.")
            else:
                exec_lines.append("No se registraron movimientos en el período seleccionado.")
            
            executive_summary("\n".join(exec_lines))
            
            # Key Metrics Dashboard
            h2("📊 Indicadores Clave")
            metric_card("Total movimientos", f"{total_movements:,}")
            metric_card("Total unidades", f"{total_quantity:,}")
            if total_cost:
                metric_card("Costo total", f"${total_cost:,.2f}")
            metric_card("Validados", f"{validated_count} ({validated_pct:.1f}%)", 
                       "good" if validated_pct >= 80 else ("warning" if validated_pct >= 50 else "critical"))
            metric_card("Pendientes", f"{pending_count} ({pending_pct:.1f}%)",
                       "good" if pending_pct == 0 else ("warning" if pending_pct < 10 else "critical"))
            metric_card("Rechazados", f"{rejected_count} ({rejected_pct:.1f}%)",
                       "good" if rejected_pct == 0 else ("warning" if rejected_pct < 5 else "critical"))
            
            # Recommendations
            recommendations = []
            if pending_count > 0:
                recommendations.append(f"Priorizar validación de {pending_count} movimientos pendientes")
            if rejected_count > 0:
                recommendations.append(f"Investigar {rejected_count} rechazos: revisar patrones (usuario, tipo, producto)")
            if validated_pct < 80 and total_movements > 0:
                recommendations.append("Implementar recordatorios automáticos para validación oportuna")
            if total_cost > 0 and avg_per_movement > 100:
                recommendations.append("Evaluar consolidación de movimientos pequeños para reducir costos transaccionales")
            
            if recommendations:
                h2("🎯 Recomendaciones")
                for i, rec in enumerate(recommendations, 1):
                    recommendation(f"{i}. {rec}")
            
            h2("Resumen")
            kv("Total movimientos:", total_movements)
            kv("Total unidades movidas:", total_quantity)
            if total_cost:
                kv("Costo total movimientos:", f"${total_cost:,.2f}")
            
            if by_type:
                h2("Por tipo")
                table(
                    ["Tipo", "Cantidad", "Unidades", "Costo total"],
                    [[t, v["count"], v["total_quantity"], f"${v.get('total_cost', 0):,.2f}"]
                      for t, v in by_type.items()],
                )
            
            if by_state:
                h2("Por estado")
                table(
                    ["Estado", "Cantidad"],
                    [[s, c] for s, c in by_state.items()],
                )
            
            if report_data.get("by_branch"):
                h2("Por sucursal")
                table(
                    ["Sucursal", "Movimientos", "Unidades", "Costo total"],
                    [[b, v["movements"], v["quantity"], f"${v.get('total_cost', 0):,.2f}"]
                      for b, v in report_data["by_branch"].items()],
                )
            
            if report_data.get("by_day_of_week"):
                h2("Por día de la semana")
                table(
                    ["Día", "Movimientos"],
                    [[d, c] for d, c in report_data["by_day_of_week"].items()],
                )
            
            h2("⚠ Puntos de Atención")
            pending = by_state.get("pendiente", 0)
            if pending > 0:
                alert(f"{pending} movimientos pendientes de validación requieren atención")
            else:
                success("No hay movimientos pendientes de validación")
            
            if report_data.get("movements"):
                h2(f"Detalle de movimientos (primeros {min(len(report_data['movements']), 100)})")
                table(
                    ["Fecha", "Producto", "SKU", "Cantidad", "Costo unit.", "Costo total", "Tipo", "Estado", "Origen", "Destino", "Usuario"],
                    [[m["date"], m["product"], m["sku"], m["quantity"],
                      f"${m.get('unit_cost', 0):,.2f}" if m.get('unit_cost') else "—",
                      f"${m.get('total_cost', 0):,.2f}" if m.get('total_cost') else "—",
                      m["type"], m["state"],
                      m["origin_branch"], m["destination_branch"], m["user"]]
                      for m in report_data["movements"][:100]],
                )
                if len(report_data["movements"]) > 100:
                    lines.append(f"\n  … y {len(report_data['movements']) - 100} movimientos más (exporta a CSV/Excel para verlos todos)")

# ── Discrepancias ─────────────────────────────────────────────────────
        elif report_type == "discrepancies":
            s = report_data.get("summary", {})
            items = report_data.get("items", [])
            total_discrepancies = report_data.get("total_discrepancies", 0)
            total_diff_abs = s.get("total_difference", 0)
            positive_diff = s.get("positive_differences", 0)
            negative_diff = s.get("negative_differences", 0)
            total_impact = sum(i.get("impact_value", 0) for i in items)
            
            # Executive Summary
            exec_lines = []
            if total_discrepancies > 0:
                exec_lines.append(f"Se detectaron {total_discrepancies} discrepancias con diferencia absoluta total de {total_diff_abs} unidades.")
                exec_lines.append(f"Desglose: {positive_diff} con exceso (físico > digital), {negative_diff} con faltante (físico < digital).")
                exec_lines.append(f"Impacto económico estimado: ${total_impact:,.2f}.")
                
                # Assess severity
                if total_impact == 0:
                    exec_lines.append("✅ Impacto económico: NULO (sin costos asociados).")
                elif total_impact < 1000:
                    exec_lines.append(f"✅ Impacto económico: BAJO (${total_impact:,.2f}).")
                elif total_impact < 10000:
                    exec_lines.append(f"⚠️  Impacto económico: MODERADO (${total_impact:,.2f}) - Revisar principales.")
                else:
                    exec_lines.append(f"🔴 Impacto económico: ALTO (${total_impact:,.2f}) - INVESTIGACIÓN URGENTE.")
                
                # Top offenders
                if items:
                    top_item = max(items, key=lambda x: x.get('impact_value', 0))
                    exec_lines.append(f"Mayor impacto: {top_item['product']} ({top_item['branch']}) - ${top_item.get('impact_value', 0):,.2f}")
            else:
                exec_lines.append("✅ No se detectaron discrepancias en el período.")
            
            executive_summary("\n".join(exec_lines))
            
            # Key Metrics
            h2("📊 Indicadores Clave")
            metric_card("Total discrepancias", f"{total_discrepancies}")
            metric_card("Diferencia absoluta", f"{total_diff_abs:,}")
            metric_card("Exceso (físico > digital)", f"{positive_diff}")
            metric_card("Faltante (físico < digital)", f"{negative_diff}")
            metric_card("Impacto económico", f"${total_impact:,.2f}",
                       "good" if total_impact == 0 else ("warning" if total_impact < 10000 else "critical"))
            
            # Recommendations
            recommendations = []
            if total_discrepancies > 0:
                high_impact = [i for i in items if i.get('impact_value', 0) > total_impact * 0.2]
                if high_impact:
                    recommendations.append(f"Investigar {len(high_impact)} discrepancias de alto impacto (>20% del total)")
                if negative_diff > positive_diff:
                    recommendations.append("Predominan faltantes: revisar procesos de recepción, picking y mermas")
                elif positive_diff > negative_diff:
                    recommendations.append("Predominan excesos: revisar recepciones no registradas y conteos duplicados")
                recommendations.append("Programar conteo cíclico en productos con discrepancias recurrentes")
            
            if recommendations:
                h2("🎯 Recomendaciones")
                for i, rec in enumerate(recommendations, 1):
                    recommendation(f"{i}. {rec}")
            
            h2("Resumen")
            kv("Total discrepancias:", total_discrepancies)
            kv("Diferencia absoluta total:", total_diff_abs)
            kv("Con exceso (físico > digital):", positive_diff)
            kv("Con faltante (físico < digital):", negative_diff)
            kv("Impacto económico total:", f"${total_impact:,.2f}")
            if items:
                h2("Detalle")
                table(
                    ["Producto", "SKU", "Sucursal", "Físico", "Digital", "Diferencia", "%", "Impacto $"],
                    [[i["product"], i["sku"], i["branch"],
                      i["physical"], i["digital"], i["difference"], i["percentage"],
                      f"${i.get('impact_value', 0):,.2f}"]
                      for i in items[:50]],
                )

        # ── KPIs ──────────────────────────────────────────────────────────────
        elif report_type == "kpis":
            kpis = report_data.get("kpis", {})
            metrics = report_data.get("metrics", {})
            
            eri = kpis.get("eri", 0)
            eru = kpis.get("eru", 0)
            
            # Determine overall health
            eri_status = "good" if eri >= 0.8 else ("warning" if eri >= 0.6 else "critical")
            eru_status = "good" if eru >= 0.85 else ("warning" if eru >= 0.7 else "critical")
            overall_good = eri_status == "good" and eru_status == "good"
            
            # Executive Summary
            exec_lines = []
            exec_lines.append("Indicadores clave de rendimiento de inventario:")
            exec_lines.append(f"• ERI (Eficiencia de Rotación): {eri:.2f} — {'Óptimo' if eri >= 0.8 else ('Aceptable' if eri >= 0.6 else 'Crítico')} (≥0.80)")
            exec_lines.append(f"• ERU (Eficiencia de Reabastecimiento): {eru:.2f} — {'Óptimo' if eru >= 0.85 else ('Aceptable' if eru >= 0.7 else 'Crítico')} (≥0.85)")
            
            if overall_good:
                exec_lines.append("✅ Salud general: EXCELENTE - Ambos KPIs en zona óptima.")
            elif eri_status == "good" or eru_status == "good":
                exec_lines.append("⚠️  Salud general: ACEPTABLE - Un KPI óptimo, el otro requiere atención.")
            else:
                exec_lines.append("🔴 Salud general: CRÍTICA - Ambos KPIs por debajo de objetivos.")
            
            # Add context from metrics
            disc_count = metrics.get("discrepancy_count", 0)
            low_stock_count = metrics.get("low_stock_count", 0)
            pending_mov = metrics.get("pending_movements", 0)
            
            if disc_count > 0:
                exec_lines.append(f"⚠️  {disc_count} discrepancias activas afectando rotación.")
            if low_stock_count > 0:
                exec_lines.append(f"⚠️  {low_stock_count} productos bajo mínimo afectando servicio.")
            if pending_mov > 0:
                exec_lines.append(f"⏳ {pending_mov} movimientos pendientes de validación.")
            
            executive_summary("\n".join(exec_lines))
            
            # Key Metrics Dashboard
            h2("📊 Dashboard de KPIs")
            metric_card("ERI (Rotación)", f"{eri:.2%}", eri_status)
            metric_card("ERU (Reabastecimiento)", f"{eru:.2%}", eru_status)
            metric_card("Discrepancias activas", f"{disc_count}", "good" if disc_count == 0 else "warning")
            metric_card("Stock bajo mínimo", f"{low_stock_count}", "good" if low_stock_count == 0 else "warning")
            metric_card("Movimientos pendientes", f"{pending_mov}", "good" if pending_mov == 0 else "warning")
            
            # Recommendations
            recommendations = []
            if eri < 0.8:
                recommendations.append("ERI bajo: Revisar productos de baja rotación, considerar liquidación o promociones")
            if eru < 0.85:
                recommendations.append("ERU bajo: Optimizar puntos de reorden y lead times de proveedores")
            if disc_count > 0:
                recommendations.append(f"Resolver {disc_count} discrepancias para mejorar precisión de KPIs")
            if low_stock_count > 0:
                recommendations.append(f"Reponer {low_stock_count} productos críticos para evitar stockouts")
            if pending_mov > 0:
                recommendations.append(f"Validar {pending_mov} movimientos pendientes para datos actualizados")
            
            if recommendations:
                h2("🎯 Plan de Acción")
                for i, rec in enumerate(recommendations, 1):
                    recommendation(f"{i}. {rec}")
            
            h2("KPIs Detalle")
            for k, v in kpis.items():
                kv(k.upper() + ":", f"{v:.2f}" if isinstance(v, float) else v)
            h2("Métricas Operativas")
            for k, v in metrics.items():
                kv(k + ":", v)

        # ── Top productos ─────────────────────────────────────────────────────
        elif report_type == "top_products":
            metric = report_data.get("metric", "—")
            limit = report_data.get("limit", "—")
            products = report_data.get("products", [])
            
            if products:
                # Executive Summary
                exec_lines = []
                exec_lines.append(f"Ranking de los Top {len(products)} productos por: {metric}.")
                
                metric_labels = {
                    "most_moved": "más movidos (volumen)",
                    "least_moved": "menos movidos (rotación lenta)",
                    "most_stock": "mayor stock (potencial exceso)",
                    "lowest_stock": "menor stock (riesgo stockout)",
                    "most_discrepancies": "más discrepancias (control calidad)"
                }
                metric_desc = metric_labels.get(metric, metric)
                exec_lines.append(f"Criterio: {metric_desc}.")
                
                if "total_quantity" in products[0]:
                    total_qty = sum(p.get("total_quantity", 0) for p in products)
                    exec_lines.append(f"Volumen combinado: {total_qty:,} unidades.")
                elif "total_stock" in products[0]:
                    total_stock = sum(p.get("total_stock", 0) for p in products)
                    exec_lines.append(f"Stock combinado: {total_stock:,} unidades.")
                elif "total_discrepancy" in products[0]:
                    total_disc = sum(p.get("total_discrepancy", 0) for p in products)
                    exec_lines.append(f"Discrepancia combinada: {total_disc:,} unidades.")
                
                executive_summary("\n".join(exec_lines))
                
                # Metrics
                h2("📊 Indicadores")
                metric_card("Métrica", metric_desc)
                metric_card("Productos en ranking", str(len(products)))
                metric_card("Límite solicitado", str(limit))
                
                # Recommendations
                recommendations = []
                if metric == "least_moved":
                    recommendations.append("Productos de rotación lenta: evaluar promociones, bundles o liquidación")
                elif metric == "lowest_stock":
                    recommendations.append("Riesgo de stockout: priorizar órdenes de compra para items críticos")
                elif metric == "most_stock":
                    recommendations.append("Exceso de inventario: analizar demanda real vs stock actual")
                elif metric == "most_discrepancies":
                    recommendations.append("Problemas de control: auditar procesos de recepción/conteo en estos SKUs")
                else:
                    recommendations.append("Productos estrella: asegurar disponibilidad continua y visibilidad en piso")
                
                if recommendations:
                    h2("🎯 Recomendaciones")
                    for i, rec in enumerate(recommendations, 1):
                        recommendation(f"{i}. {rec}")
                
                h2("Ranking Detallado")
                first = products[0]
                hdrs = list(first.keys())
                table(hdrs, [[p.get(h) for h in hdrs] for p in products])
            else:
                lines.append("    No hay datos para mostrar con los filtros actuales.")

# ── Eficiencia por sucursal ────────────────────────────────────────────
        elif report_type == "branch_efficiency":
            branches = report_data.get("branches", [])
            if branches:
                high_rejection = [b for b in branches if b.get("rejection_rate", 0) > 15]
                high_discrepancy = [b for b in branches if b.get("discrepancy_rate", 0) > 10]
                slow_validation = [b for b in branches if b.get("avg_validation_days", 0) > 3]
                total_branches = len(branches)
                healthy_branches = total_branches - len(high_rejection) - len(high_discrepancy) - len(slow_validation)
                
                # Executive Summary
                exec_lines = []
                exec_lines.append(f"Análisis de eficiencia operativa para {total_branches} sucursales.")
                exec_lines.append(f"Sucursales saludables: {healthy_branches}/{total_branches} ({healthy_branches/total_branches*100:.0f}%).")
                
                if high_rejection:
                    exec_lines.append(f"🔴 {len(high_rejection)} con ALTA TASA DE RECHAZO (>15%): {', '.join(b['branch'] for b in high_rejection)}")
                if high_discrepancy:
                    exec_lines.append(f"🔴 {len(high_discrepancy)} con ALTA DISCREPANCIA (>10%): {', '.join(b['branch'] for b in high_discrepancy)}")
                if slow_validation:
                    exec_lines.append(f"⚠️  {len(slow_validation)} con VALIDACIÓN LENTA (>3 días): {', '.join(b['branch'] for b in slow_validation)}")
                
                if healthy_branches == total_branches:
                    exec_lines.append("✅ TODAS las sucursales operan dentro de parámetros aceptables.")
                
                executive_summary("\n".join(exec_lines))
                
                # Metrics
                h2("📊 Scorecard por Sucursal")
                metric_card("Total sucursales", str(total_branches))
                metric_card("Saludables", f"{healthy_branches}/{total_branches}")
                metric_card("Alta rechazo", str(len(high_rejection)), "critical" if high_rejection else "good")
                metric_card("Alta discrepancia", str(len(high_discrepancy)), "critical" if high_discrepancy else "good")
                metric_card("Validación lenta", str(len(slow_validation)), "warning" if slow_validation else "good")
                
                # Recommendations
                recommendations = []
                if high_rejection:
                    recommendations.append(f"Capacitar equipos en {len(high_rejection)} sucursales con alta tasa de rechazo")
                if high_discrepancy:
                    recommendations.append(f"Auditar procesos de conteo/recepción en {len(high_discrepancy)} sucursales problemáticas")
                if slow_validation:
                    recommendations.append(f"Implementar SLA de validación 48h en {len(slow_validation)} sucursales lentas")
                if healthy_branches == total_branches:
                    recommendations.append("Mantener estándares actuales; compartir mejores prácticas entre sucursales")
                
                if recommendations:
                    h2("🎯 Plan de Mejora")
                    for i, rec in enumerate(recommendations, 1):
                        recommendation(f"{i}. {rec}")
                
                h2("Eficiencia por Sucursal")
                table(
                    ["Sucursal", "Movimientos", "Tasa rechazo %",
                     "Días prom. validación", "Transf. enviadas", "Transf. recibidas",
                     "Tasa discrepancias %", "Estado"],
                    [[b["branch"], b["total_movements"], f"{b['rejection_rate']:.1f}%",
                      f"{b['avg_validation_days']:.1f}" if b['avg_validation_days'] else "—",
                      b["transfers_sent"], b["transfers_received"], f"{b['discrepancy_rate']:.1f}%",
                      "🔴 Crítica" if b.get("rejection_rate", 0) > 15 or b.get("discrepancy_rate", 0) > 10 
                      else ("🟡 Atención" if b.get("avg_validation_days", 0) > 3 else "🟢 OK")]
                      for b in branches],
                )

        # ── Valor de inventario ───────────────────────────────────────────────
        elif report_type == "inventory_value":
            total_value = report_data.get("total_value", 0)
            total_items = report_data.get("total_items", 0)
            no_price_count = report_data.get("no_price_count", 0)
            by_branch = report_data.get("by_branch", {})
            items = report_data.get("items", [])
            
            avg_value = (total_value / total_items) if total_items > 0 else 0
            pct_no_price = (no_price_count / total_items * 100) if total_items > 0 else 0
            
            # Executive Summary
            exec_lines = []
            exec_lines.append(f"Valor total de inventario: ${total_value:,.2f} distribuido en {total_items:,} artículos.")
            exec_lines.append(f"Valor promedio por artículo: ${avg_value:,.2f}.")
            
            if no_price_count > 0:
                exec_lines.append(f"⚠️  {no_price_count} artículos ({pct_no_price:.1f}%) SIN precio asignado - valor = $0.")
                exec_lines.append("   Estos items no contribuyen al valor total; asignar costos para precisión.")
            else:
                exec_lines.append("✅ Todos los artículos tienen precio/costo asignado.")
            
            # Top concentration
            if items:
                top5_value = sum(i.get("value", 0) for i in items[:5])
                top5_pct = (top5_value / total_value * 100) if total_value > 0 else 0
                exec_lines.append(f"Top 5 items concentran {top5_pct:.1f}% del valor total (${top5_value:,.2f}).")
            
            executive_summary("\n".join(exec_lines))
            
            # Key Metrics
            h2("📊 Indicadores Clave")
            metric_card("Valor total inventario", f"${total_value:,.2f}")
            metric_card("Total artículos", f"{total_items:,}")
            metric_card("Valor promedio/item", f"${avg_value:,.2f}")
            metric_card("Sin precio asignado", f"{no_price_count} ({pct_no_price:.1f}%)",
                       "good" if no_price_count == 0 else "warning")
            
            # Branch distribution
            if by_branch:
                h2("Distribución por Sucursal")
                table(
                    ["Sucursal", "Artículos", "Valor Total", "% del Total"],
                    [[b, v["items"], f"${v['value']:,.2f}", f"{(v['value']/total_value*100):.1f}%" if total_value > 0 else "0%"]
                     for b, v in sorted(by_branch.items(), key=lambda x: x[1]["value"], reverse=True)],
                )
            
            # Recommendations
            recommendations = []
            if no_price_count > 0:
                recommendations.append(f"Asignar costos a {no_price_count} productos sin precio (revisar unit_cost en Inventory o unit_price en Product)")
            if items:
                zero_value = [i for i in items if i.get("value", 0) == 0]
                if zero_value:
                    recommendations.append(f"{len(zero_value)} productos con valor $0: revisar si stock real o error de datos")
            if total_value > 1000000:
                recommendations.append("Inventario >$1M: verificar cobertura de seguro y límites de riesgo")
            
            if recommendations:
                h2("🎯 Recomendaciones")
                for i, rec in enumerate(recommendations, 1):
                    recommendation(f"{i}. {rec}")
            
            # Detail
            if items:
                h2("Detalle (Top 50 por valor)")
                table(
                    ["Producto", "SKU", "Sucursal", "Stock", "Costo unit.", "Valor", "% Total"],
                    [[i["product"], i["sku"], i["branch"],
                      i["stock"], i["unit_price"] if i["unit_price"] else "—", f"${i['value']:,.2f}",
                      f"{(i['value']/total_value*100):.1f}%" if total_value > 0 else "0%"]
                     for i in items[:50]],
                )

        # ── Transferencias ────────────────────────────────────────────────────
        elif report_type == "transfers":
            total_transfers = report_data.get("total_transfers", 0)
            pending_reception = report_data.get("pending_reception", 0)
            items = report_data.get("items", [])
            pending_pct = (pending_reception / total_transfers * 100) if total_transfers > 0 else 0
            
            # Executive Summary
            exec_lines = []
            if total_transfers > 0:
                exec_lines.append(f"Total de transferencias inter-sucursal: {total_transfers}.")
                exec_lines.append(f"Pendientes de recepción: {pending_reception} ({pending_pct:.1f}%).")
                
                if pending_reception == 0:
                    exec_lines.append("✅ Recepción: TODAS las transferencias han sido recibidas.")
                elif pending_pct < 10:
                    exec_lines.append(f"⚠️  {pending_reception} transferencias pendientes ({pending_pct:.1f}%) - Seguimiento rutinario.")
                else:
                    exec_lines.append(f"🔴 {pending_reception} transferencias pendientes ({pending_pct:.1f}%) - SEGUIMIENTO URGENTE.")
                
                # Analyze by branch pair
                by_pair = report_data.get("by_branch_pair", {})
                if by_pair:
                    top_pair = max(by_pair.items(), key=lambda x: x[1]["count"])
                    exec_lines.append(f"Ruta más activa: {top_pair[0]} ({top_pair[1]['count']} transferencias, {top_pair[1]['total_quantity']} unidades).")
            else:
                exec_lines.append("No hay transferencias en el período seleccionado.")
            
            executive_summary("\n".join(exec_lines))
            
            # Metrics
            h2("📊 Indicadores")
            metric_card("Total transferencias", str(total_transfers))
            metric_card("Pendientes recepción", f"{pending_reception} ({pending_pct:.1f}%)", 
                       "good" if pending_reception == 0 else ("warning" if pending_pct < 10 else "critical"))
            metric_card("Completadas", str(total_transfers - pending_reception))
            
            # Recommendations
            recommendations = []
            if pending_reception > 0:
                old_pending = [i for i in items if not i.get("is_received") and i.get("created_at")]
                if old_pending:
                    recommendations.append(f"Contactar destino para {len(old_pending)} transferencias sin confirmar recepción")
            if total_transfers > 0 and pending_pct > 20:
                recommendations.append("Alta tasa de pendientes: revisar proceso de confirmación de recepción en destino")
            if by_pair:
                unbalanced = [(k, v) for k, v in by_pair.items() if v["count"] > total_transfers * 0.3]
                if unbalanced:
                    recommendations.append(f"Ruta concentrada: {unbalanced[0][0]} - evaluar envíos directos vs consolidados")
            
            if recommendations:
                h2("🎯 Recomendaciones")
                for i, rec in enumerate(recommendations, 1):
                    recommendation(f"{i}. {rec}")
            
            # Detail
            if items:
                h2(f"Detalle (primeras 50 de {len(items)})")
                table(
                    ["Producto", "Origen", "Destino", "Cantidad", "Estado", "Recibida", "Fecha"],
                    [[i["product"], i["origin_branch"], i["destination_branch"],
                      i["quantity"], i["state"], "Sí" if i["is_received"] else "No",
                      i.get("created_at", "—")[:10]]
                     for i in items[:50]],
                )

        # ── Tendencias ────────────────────────────────────────────────────────
        elif report_type == "trends":
            metric = report_data.get("metric", "—")
            periods_count = report_data.get("periods_count", "—")
            data_points = report_data.get("data_points", [])
            trend = report_data.get("trend", "—")
            
            metric_labels = {
                "movimientos_total": "Total de movimientos",
                "stock_total": "Stock total",
                "discrepancias_count": "Conteo de discrepancias"
            }
            metric_desc = metric_labels.get(metric, metric)
            
            if data_points and len(data_points) >= 2:
                first_val = data_points[0].get("value", 0)
                last_val = data_points[-1].get("value", 0)
                change = last_val - first_val
                change_pct = (change / first_val * 100) if first_val != 0 else 0
                
                # Determine trend direction
                if change > 0:
                    trend_icon = "📈"
                    trend_word = "CRECIENTE"
                elif change < 0:
                    trend_icon = "📉"
                    trend_word = "DECRECIENTE"
                else:
                    trend_icon = "➡️"
                    trend_word = "ESTABLE"
                
                # Executive Summary
                exec_lines = []
                exec_lines.append(f"Análisis de tendencia para: {metric_desc} ({periods_count} períodos).")
                exec_lines.append(f"Dirección general: {trend_icon} {trend_word} ({change_pct:+.1f}%).")
                exec_lines.append(f"Valor inicial: {first_val:,} → Valor final: {last_val:,} (Δ {change:+,}).")
                
                # Assess significance
                if abs(change_pct) > 20:
                    exec_lines.append(f"⚠️  Variación SIGNIFICATIVA del {abs(change_pct):.1f}% - Investigar causas raíz.")
                elif abs(change_pct) > 10:
                    exec_lines.append(f"📊 Variación MODERADA del {abs(change_pct):.1f}% - Monitorear evolución.")
                else:
                    exec_lines.append(f"✅ Variación MENOR del {abs(change_pct):.1f}% - Dentro de rango normal.")
                
                # Peak/valley
                max_val = max(p.get("value", 0) for p in data_points)
                min_val = min(p.get("value", 0) for p in data_points)
                max_period = next(p["period"] for p in data_points if p.get("value", 0) == max_val)
                min_period = next(p["period"] for p in data_points if p.get("value", 0) == min_val)
                exec_lines.append(f"Pico: {max_val:,} ({max_period}) | Valle: {min_val:,} ({min_period})")
                
                executive_summary("\n".join(exec_lines))
                
                # Metrics
                h2("📊 Indicadores de Tendencia")
                metric_card("Métrica", metric_desc)
                metric_card("Períodos analizados", str(periods_count))
                metric_card("Dirección", f"{trend_icon} {trend_word} ({change_pct:+.1f}%)",
                           "good" if change > 0 else ("warning" if change == 0 else "critical"))
                metric_card("Cambio absoluto", f"{change:+,}")
                metric_card("Rango (min-max)", f"{min_val:,} - {max_val:,}")
                
                # Recommendations
                recommendations = []
                if abs(change_pct) > 20:
                    recommendations.append(f"Variación >20%: Investigar eventos externos (promociones, proveedores, estacionalidad)")
                if change < 0 and metric == "movimientos_total":
                    recommendations.append("Caída en movimientos: revisar demanda, competencia y calendario comercial")
                if change > 0 and metric == "discrepancias_count":
                    recommendations.append("Aumento de discrepancias: auditar procesos de conteo y recepción")
                if abs(change_pct) < 5:
                    recommendations.append("Tendencia estable: mantener controles actuales, planificar capacidad futura")
                
                if recommendations:
                    h2("🎯 Recomendaciones")
                    for i, rec in enumerate(recommendations, 1):
                        recommendation(f"{i}. {rec}")
            
            if data_points:
                h2("Serie Temporal")
                table(
                    ["Período", "Valor", "Δ vs anterior"],
                    [[p["period"], f"{p['value']:,}", 
                      f"{(p['value'] - data_points[i-1]['value']):+,}" if i > 0 else "—"]
                     for i, p in enumerate(data_points)],
                )

        # ── Comparación ───────────────────────────────────────────────────────
        elif report_type == "comparison":
            p1 = report_data.get("period1", {})
            p2 = report_data.get("period2", {})
            diffs = report_data.get("differences", {})
            report_type_name = report_data.get("report_type", "").replace("_", " ").title()
            
            # Executive Summary
            exec_lines = []
            exec_lines.append(f"Comparación de {report_type_name} entre dos períodos.")
            
            p1_from = report_data.get("period1", {}).get("from", "—")[:10]
            p1_to = report_data.get("period1", {}).get("to", "—")[:10]
            p2_from = report_data.get("period2", {}).get("from", "—")[:10]
            p2_to = report_data.get("period2", {}).get("to", "—")[:10]
            exec_lines.append(f"Período 1: {p1_from} a {p1_to}")
            exec_lines.append(f"Período 2: {p2_from} a {p2_to}")
            
            # Key differences
            significant_changes = []
            for k, v in diffs.items():
                if isinstance(v, dict) and "delta_pct" in v:
                    pct = v["delta_pct"]
                    if abs(pct) > 10:
                        direction = "📈" if pct > 0 else "📉"
                        significant_changes.append(f"{direction} {k}: {v['period1']:,} → {v['period2']:,} ({pct:+.1f}%)")
            
            if significant_changes:
                exec_lines.append("Cambios significativos (>10%):")
                for ch in significant_changes[:5]:
                    exec_lines.append(f"  • {ch}")
            else:
                exec_lines.append("✅ Sin cambios significativos (>10%) entre períodos.")
            
            executive_summary("\n".join(exec_lines))
            
            # Metrics
            h2("📊 Resumen Comparativo")
            metric_card("Tipo de reporte", report_type_name)
            metric_card("Período 1", f"{p1_from} a {p1_to}")
            metric_card("Período 2", f"{p2_from} a {p2_to}")
            metric_card("Métricas con cambio >10%", str(len(significant_changes)))
            
            # Detailed comparison
            for period_key, period_label in [("period1", "Período 1"), ("period2", "Período 2")]:
                p = report_data.get(period_key, {})
                if p and "data" in p:
                    h2(period_label)
                    data = p["data"]
                    for k, v in data.items():
                        if not isinstance(v, (dict, list)):
                            kv(k + ":", v)
            
            if diffs:
                h2("Diferencias Detalladas")
                table(
                    ["Métrica", "Período 1", "Período 2", "Δ Absoluto", "Δ %"],
                    [[k, f"{v['period1']:,}", f"{v['period2']:,}", f"{v['delta']:+,}", f"{v['delta_pct']:+.1f}%"]
                     for k, v in diffs.items() if isinstance(v, dict) and "delta_pct" in v],
                )
            
            # Recommendations
            if significant_changes:
                h2("🎯 Recomendaciones")
                recommendation("1. Analizar causas de variaciones >10% (estacionalidad, promociones, problemas operativos)")
                recommendation("2. Comparar con mismo período año anterior para detectar patrones")
                recommendation("3. Ajustar pronósticos y planes de compra según nueva tendencia")

        # ── Actividad usuario ─────────────────────────────────────────────────
        elif report_type == "user_activity":
            total_activities = report_data.get("total_activities", 0)
            by_event = report_data.get("by_event_type", {})
            activities = report_data.get("activities", [])
            
            # Executive Summary
            exec_lines = []
            exec_lines.append(f"Total de actividades registradas: {total_activities:,}.")
            
            if by_event:
                top_event = max(by_event.items(), key=lambda x: x[1])
                exec_lines.append(f"Evento más frecuente: {top_event[0]} ({top_event[1]:,} veces, {top_event[1]/total_activities*100:.1f}%).")
                exec_lines.append(f"Tipos de eventos distintos: {len(by_event)}.")
            
            if activities:
                unique_users = len(set(a.get("user", "—") for a in activities))
                exec_lines.append(f"Usuarios activos: {unique_users}.")
                
                # Recent activity
                exec_lines.append(f"Período cubre {len(set(a.get('fecha', '—')[:10] for a in activities))} días de actividad.")
            
            executive_summary("\n".join(exec_lines))
            
            # Metrics
            h2("📊 Indicadores de Actividad")
            metric_card("Total actividades", f"{total_activities:,}")
            metric_card("Tipos de evento", str(len(by_event)))
            if activities:
                unique_users = len(set(a.get("user", "—") for a in activities))
                metric_card("Usuarios activos", str(unique_users))
            
            if by_event:
                h2("Por tipo de evento")
                table(
                    ["Evento", "Cantidad", "%"],
                    [[e, f"{c:,}", f"{c/total_activities*100:.1f}%"] for e, c in sorted(by_event.items(), key=lambda x: -x[1])],
                )

        # ── Auditoría de Historial ────────────────────────────────────────────
        elif report_type == "history_audit":
            total_entries = report_data.get("total_entries", 0)
            returned = report_data.get("returned_entries", 0)
            by_entity = report_data.get("by_entity_type", {})
            by_event = report_data.get("by_event_type", {})
            by_user = report_data.get("by_user", {})
            by_date = report_data.get("by_date", {})
            top_entities = report_data.get("top_entities", [])
            items = report_data.get("items", [])
            
            # Executive Summary
            exec_lines = []
            exec_lines.append(f"Auditoría de historial: {returned:,} de {total_entries:,} registros analizados ({returned/total_entries*100:.1f}%)." if total_entries > 0 else "Sin registros en el período.")
            
            if by_entity:
                top_ent = max(by_entity.items(), key=lambda x: x[1])
                exec_lines.append(f"Entidad más auditada: {top_ent[0]} ({top_ent[1]:,} eventos, {top_ent[1]/returned*100:.1f}%).")
            
            if by_event:
                top_evt = max(by_event.items(), key=lambda x: x[1])
                exec_lines.append(f"Evento más frecuente: {top_evt[0]} ({top_evt[1]:,} veces).")
            
            if by_user:
                top_usr = max(by_user.items(), key=lambda x: x[1])
                exec_lines.append(f"Usuario más activo: {top_usr[0]} ({top_usr[1]:,} acciones).")
            
            if by_date:
                exec_lines.append(f"Período cubre {len(by_date)} días de actividad.")
            
            # Detect anomalies
            deleted_count = sum(1 for i in items if i.get("eliminado"))
            if deleted_count > 0:
                exec_lines.append(f"⚠️  {deleted_count} registros de entidades eliminadas detectados.")
            
            executive_summary("\n".join(exec_lines))
            
            # Metrics
            h2("📊 Indicadores de Auditoría")
            metric_card("Total en BD", f"{total_entries:,}")
            metric_card("En reporte", f"{returned:,}")
            metric_card("Tipos de entidad", str(len(by_entity)))
            metric_card("Tipos de evento", str(len(by_event)))
            metric_card("Usuarios activos", str(len(by_user)))
            metric_card("Días con actividad", str(len(by_date)))
            if items:
                metric_card("Entidades eliminadas", str(deleted_count), "warning" if deleted_count > 0 else "good")
            
            # Tables
            if by_entity:
                h2("Por tipo de entidad")
                table(["Tipo", "Eventos", "%"], [[t, f"{c:,}", f"{c/returned*100:.1f}%"] for t, c in sorted(by_entity.items(), key=lambda x: -x[1])])
            
            if by_event:
                h2("Por tipo de evento (Top 20)")
                table(["Evento", "Cantidad", "%"], [[e, f"{c:,}", f"{c/returned*100:.1f}%"] for e, c in list(sorted(by_event.items(), key=lambda x: -x[1]))[:20]])
            
            if by_user:
                h2("Por usuario")
                table(["Usuario", "Acciones", "%"], [[u, f"{c:,}", f"{c/returned*100:.1f}%"] for u, c in sorted(by_user.items(), key=lambda x: -x[1])])
            
            if by_date:
                h2("Actividad diaria")
                table(["Fecha", "Eventos"], [[_fmt_date_iso_to_readable(d), f"{c:,}"] for d, c in sorted(by_date.items())])
            
            if top_entities:
                h2("Entidades más activas (Top 20)")
                table(["Tipo", "Entidad", "Eventos"], [[e["entity_type"], e["entity_name"], f"{e['event_count']:,}"] for e in top_entities])
            
            if items:
                h2(f"Detalle ({len(items)} registros)")
                table(
                    ["ID", "Fecha", "Evento", "Tipo", "Entidad", "Usuario", "Acción", "Eliminado"],
                    [[i["id"], _fmt_date_iso_to_readable(i["fecha"]), i["evento"], i["tipo"],
                      i["entidad"], i["usuario"], i["accion"], "Sí" if i.get("eliminado") else "No"]
                     for i in items[:200]],
                )
                if len(items) > 200:
                    lines.append(f"\n  … y {len(items) - 200} registros más (exporta a CSV/Excel para verlos todos)")

        elif report_type == "count_sessions":
            s = report_data.get("summary", {})
            total_sessions = s.get("total_sessions", 0)
            completed = s.get("completed", 0)
            in_progress = s.get("in_progress", 0)
            pending = s.get("pending", 0)
            items_counted = s.get("total_items_counted", 0)
            discrepancies = s.get("total_discrepancies_found", 0)
            completion_rate = (completed / total_sessions * 100) if total_sessions > 0 else 0
            
            # Executive Summary
            exec_lines = []
            if total_sessions > 0:
                exec_lines.append(f"Programadas {total_sessions} sesiones de conteo: {completed} completadas ({completion_rate:.1f}%), {in_progress} en progreso, {pending} pendientes.")
                exec_lines.append(f"Total ítems contados: {items_counted:,} | Discrepancias detectadas: {discrepancies}.")
                
                if completion_rate == 100:
                    exec_lines.append("✅ Completitud: TODAS las sesiones finalizadas.")
                elif completion_rate >= 80:
                    exec_lines.append(f"⚠️  Completitud {completion_rate:.1f}% - {in_progress} en progreso requieren cierre.")
                else:
                    exec_lines.append(f"🔴 Completitud BAJA ({completion_rate:.1f}%) - {pending} pendientes + {in_progress} en curso.")
                
                if discrepancies > 0:
                    exec_lines.append(f"⚠️  {discrepancies} discrepancias halladas - Revisar ajustes necesarios.")
            else:
                exec_lines.append("No hay sesiones de conteo en el período.")
            
            executive_summary("\n".join(exec_lines))
            
            # Metrics
            h2("📊 Scorecard Conteo")
            metric_card("Total sesiones", str(total_sessions))
            metric_card("Completadas", f"{completed} ({completion_rate:.1f}%)", "good" if completion_rate == 100 else ("warning" if completion_rate >= 80 else "critical"))
            metric_card("En progreso", str(in_progress), "warning" if in_progress > 0 else "good")
            metric_card("Pendientes", str(pending), "good" if pending == 0 else "warning")
            metric_card("Ítems contados", f"{items_counted:,}")
            metric_card("Discrepancias", str(discrepancies), "good" if discrepancies == 0 else "warning")
            
            # Recommendations
            recommendations = []
            if pending > 0:
                recommendations.append(f"Programar inicio de {pending} sesiones pendientes")
            if in_progress > 0:
                recommendations.append(f"Dar seguimiento a {in_progress} sesiones en progreso para cierre oportuno")
            if discrepancies > 0:
                recommendations.append("Investigar discrepancias y aplicar ajustes de inventario")
            if completion_rate < 80 and total_sessions > 0:
                recommendations.append("Revisar proceso: baja completitud sugiere problemas de planificación o ejecución")
            
            if recommendations:
                h2("🎯 Recomendaciones")
                for i, rec in enumerate(recommendations, 1):
                    recommendation(f"{i}. {rec}")
            
            if report_data.get("sessions"):
                h2("Detalle de Sesiones")
                table(
                    ["ID", "Sucursal", "Fecha", "Estado", "Validador", "Ítems", "Discrepancias"],
                    [[srow["session_id"], srow["branch_name"], _fmt_date_iso_to_readable(srow["scheduled_date"]), srow["status"], srow["validator_count"], srow["items_total"], srow["items_with_discrepancy"]] for srow in report_data["sessions"][:50]],
                )

        elif report_type == "approvals":
            s = report_data.get("summary", {})
            total_requests = s.get("total_transfer_requests", 0)
            pending_admin = s.get("pending_admin", 0)
            pending_manager = s.get("pending_manager", 0)
            total_pending = pending_admin + pending_manager
            
            h2("Resumen")
            kv("Solicitudes totales:", total_requests)
            kv("Pendientes admin:", pending_admin)
            kv("Pendientes manager:", pending_manager)
            kv("Aprobadas:", s.get("approved", 0))
            kv("Rechazadas:", s.get("rejected", 0))
            kv("Promedio horas aprobación:", s.get("average_approval_time_hours", 0))
            
            if total_requests > 0:
                h2("Resumen Ejecutivo")
                pending_pct = (total_pending / total_requests * 100) if total_requests > 0 else 0
                avg_hours = s.get("average_approval_time_hours", 0)
                
                exec_lines = []
                exec_lines.append(f"Solicitudes de transferencia: {total_requests} totales, {total_pending} pendientes ({pending_pct:.1f}%).")
                exec_lines.append(f"Aprobadas: {s.get('approved', 0)} | Rechazadas: {s.get('rejected', 0)}.")
                exec_lines.append(f"Tiempo promedio de aprobación: {avg_hours:.1f} horas.")
                
                if total_pending == 0:
                    exec_lines.append("✅ Flujo de aprobaciones: SIN CUERLOS - todo al día.")
                elif pending_pct < 10:
                    exec_lines.append(f"⚠️  {total_pending} pendientes ({pending_pct:.1f}%) - Gestión rutinaria.")
                else:
                    exec_lines.append(f"🔴 CUELLO DE BOTELLA: {total_pending} pendientes ({pending_pct:.1f}%) - ESCALAR.")
                
                # Aging analysis
                if report_data.get("pending_approvals"):
                    old_pending = [p for p in report_data["pending_approvals"] if p.get("waiting_hours", 0) > 48]
                    if old_pending:
                        exec_lines.append(f"⏰ {len(old_pending)} solicitudes >48h sin resolver - Riesgo operativo.")
                
                executive_summary("\n".join(exec_lines))
                
                # Metrics
                h2("📊 KPIs de Aprobación")
                metric_card("Total solicitudes", str(total_requests))
                metric_card("Pendientes", f"{total_pending} ({pending_pct:.1f}%)", 
                           "good" if total_pending == 0 else ("warning" if pending_pct < 10 else "critical"))
                metric_card("Pendientes Admin", str(s.get("pending_admin", 0)))
                metric_card("Pendientes Manager", str(s.get("pending_manager", 0)))
                metric_card("Tiempo prom. aprobación", f"{avg_hours:.1f}h",
                           "good" if avg_hours < 24 else ("warning" if avg_hours < 48 else "critical"))
                
                # Recommendations
                recommendations = []
                if total_pending > 0:
                    recommendations.append("Asignar aprobador designado por turno para reducir tiempo de espera")
                    if avg_hours > 24:
                        recommendations.append("Implementar recordatorios automáticos cada 12h para aprobadores")
                if s.get("rejected", 0) > 0:
                    recommendations.append(f"Analizar {s.get('rejected', 0)} rechazos: ¿patrón de errores en origen?")
                
                if recommendations:
                    h2("🎯 Recomendaciones")
                    for i, rec in enumerate(recommendations, 1):
                        recommendation(f"{i}. {rec}")
            
            if report_data.get("pending_approvals"):
                h2("Pendientes de aprobación")
                table(
                    ["Movimiento", "Producto", "Origen", "Destino", "Cant.", "Nivel", "Creada", "Horas espera", "Urgencia"],
                    [[row["movement_id"], row["product_name"], row["origin_branch"], row["destination_branch"], row["quantity"], row["approval_level"], _fmt_date_iso_to_readable(row["created_at"]), row["waiting_hours"], "🔴 >48h" if row.get("waiting_hours", 0) > 48 else ("🟡 >24h" if row.get("waiting_hours", 0) > 24 else "🟢 <24h")] for row in report_data["pending_approvals"][:20]],
                )

        elif report_type == "alerts":
            s = report_data.get("summary", {})
            total_alerts = s.get("total_alerts", 0)
            open_alerts = s.get("open_alerts", 0)
            resolved_alerts = s.get("resolved_alerts", 0)
            avg_resolution = s.get("average_resolution_hours", 0)
            by_severity = s.get("by_severity", {})
            by_type = s.get("by_type", {})
            by_branch = report_data.get("by_branch", {})
            top_open = report_data.get("top_open_alerts", [])
            open_pct = (open_alerts / total_alerts * 100) if total_alerts > 0 else 0
            critical_open = sum(1 for a in top_open if a.get("severity", "").lower() == "critical")
            
            # Executive Summary
            exec_lines = []
            if total_alerts > 0:
                exec_lines.append(f"Total alertas en período: {total_alerts:,} | Abiertas: {open_alerts} ({open_pct:.1f}%) | Resueltas: {resolved_alerts:,}.")
                exec_lines.append(f"Tiempo promedio de resolución: {avg_resolution:.1f}h.")
                
                if open_alerts == 0:
                    exec_lines.append("✅ Gestión de alertas: LIMPIA - Sin backlog.")
                elif open_pct < 20:
                    exec_lines.append(f"⚠️  Backlog controlado: {open_alerts} abiertas - Seguimiento rutinario.")
                else:
                    exec_lines.append(f"🔴 BACKLOG ALTO: {open_alerts} abiertas ({open_pct:.1f}%) - REQUIERE ACCIÓN INMEDIATA.")
                
                if critical_open > 0:
                    exec_lines.append(f"🚨 {critical_open} alertas CRÍTICAS sin resolver - Escalar a gerencia YA.")
                
                # By branch
                if by_branch:
                    worst = max(by_branch.items(), key=lambda x: x[1].get("open", 0))
                    if worst[1].get("open", 0) > 0:
                        exec_lines.append(f"Sucursal más afectada: {worst[0]} ({worst[1]['open']} abiertas, {worst[1].get('critical_open', 0)} críticas).")
            else:
                exec_lines.append("✅ Período sin alertas - Operación normal.")
            
            executive_summary("\n".join(exec_lines))
            
            # Metrics Dashboard
            h2("📊 Panel de Control de Alertas")
            metric_card("Total alertas", f"{total_alerts:,}")
            metric_card("Abiertas", f"{open_alerts} ({open_pct:.1f}%)", 
                       "good" if open_alerts == 0 else ("warning" if open_pct < 20 else "critical"))
            metric_card("Resueltas", f"{resolved_alerts:,}")
            metric_card("Tiempo prom. resolución", f"{avg_resolution:.1f}h",
                       "good" if avg_resolution < 24 else ("warning" if avg_resolution < 48 else "critical"))
            metric_card("Críticas abiertas", str(critical_open), "critical" if critical_open > 0 else "good")
            
            # By severity
            if by_severity:
                h2("Por Severidad")
                table(["Severidad", "Cantidad", "%"], [[k, v, f"{v/total_alerts*100:.1f}%"] for k, v in sorted(by_severity.items(), key=lambda x: -x[1])])
            
            # By type
            if by_type:
                h2("Por Tipo")
                table(["Tipo", "Cantidad", "%"], [[k, v, f"{v/total_alerts*100:.1f}%"] for k, v in sorted(by_type.items(), key=lambda x: -x[1])])
            
            # By branch
            if by_branch:
                h2("Por Sucursal")
                table(["Sucursal", "Total", "Abiertas", "Críticas abiertas"], [[b, v["total"], v.get("open", 0), v.get("critical_open", 0)] for b, v in sorted(by_branch.items(), key=lambda x: -x[1].get("open", 0))])
            
            # Recommendations
            recommendations = []
            if critical_open > 0:
                recommendations.append(f"Resolver {critical_open} alertas CRÍTICAS ahora - asignar responsable y deadline 2h")
            if open_alerts > 10:
                recommendations.append(f"Backlog de {open_alerts} alertas: dedicar 30min/día a cierre masivo")
            if avg_resolution > 48:
                recommendations.append("Tiempo de resolución >48h: revisar proceso de escalamiento y recursos")
            if by_branch:
                high_branch = [(b, v) for b, v in by_branch.items() if v.get("open", 0) > 5]
                if high_branch:
                    recommendations.append(f"Sucursales con >5 abiertas: {', '.join(b for b, v in high_branch)} - auditoría local")
            
            if recommendations:
                h2("🎯 Plan de Acción Inmediato")
                for i, rec in enumerate(recommendations, 1):
                    recommendation(f"{i}. {rec}")
            
            # Top open alerts detail
            if top_open:
                h2("🔴 Alertas Abiertas Prioritarias (Top 10)")
                for row in top_open[:10]:
                    severity = row.get("severity", "info").lower()
                    icon = "🔴" if severity == "critical" else ("🟡" if severity == "high" else "🔵")
                    lines.append(f"  {icon} {row['type']} - {row['product_name']} en {row['branch_name']} ({row['hours_open']}h abierta)")
                
                h2("Detalle Top 10")
                table(
                    ["ID", "Tipo", "Severidad", "Producto", "Sucursal", "Horas abiertas"],
                    [[row["alert_id"], row["type"], row["severity"], row["product_name"], row["branch_name"], row["hours_open"]] for row in top_open[:10]],
                )

        elif report_type == "branch_capacity":
            branches = report_data.get("branches", [])
            suggestions = report_data.get("transfer_suggestions", [])
            summary = report_data.get("summary", {})
            
            total_branches = summary.get("total_branches", len(branches))
            with_limit = summary.get("with_capacity_limit", 0)
            without_limit = summary.get("without_capacity_limit", 0)
            avg_usage = summary.get("avg_usage_percent", 0)
            
            critical_branches = [b for b in branches if b.get("status") == "critical"]
            warning_branches = [b for b in branches if b.get("status") == "warning"]
            ok_branches = [b for b in branches if b.get("status") == "ok"]
            no_limit_branches = [b for b in branches if b.get("status") == "no_limit"]
            
            # Executive Summary
            exec_lines = []
            exec_lines.append(f"Análisis de capacidad para {total_branches} sucursales ({with_limit} con límite, {without_limit} sin límite).")
            exec_lines.append(f"Uso promedio: {avg_usage:.1f}%.")
            
            if critical_branches:
                exec_lines.append(f"🔴 {len(critical_branches)} CRÍTICAS (>{90}% uso): {', '.join(b['branch_name'] for b in critical_branches[:5])}.")
            if warning_branches:
                exec_lines.append(f"⚠️  {len(warning_branches)} EN ALERTA (70-90%): {', '.join(b['branch_name'] for b in warning_branches[:5])}.")
            if ok_branches:
                exec_lines.append(f"✅ {len(ok_branches)} normales (<70%): {', '.join(b['branch_name'] for b in ok_branches[:5])}.")
            if no_limit_branches:
                exec_lines.append(f"ℹ️  {len(no_limit_branches)} sin límite configurado.")
            
            executive_summary("\n".join(exec_lines))
            
            # Metrics
            h2("📊 Scorecard Capacidad")
            metric_card("Total sucursales", str(total_branches))
            metric_card("Con límite", str(with_limit))
            metric_card("Críticas (>90%)", str(len(critical_branches)), "critical" if critical_branches else "good")
            metric_card("Advertencia (70-90%)", str(len(warning_branches)), "warning" if warning_branches else "good")
            metric_card("OK (<70%)", str(len(ok_branches)), "good")
            metric_card("Sin límite", str(len(no_limit_branches)))
            metric_card("Uso promedio", f"{avg_usage:.1f}%")
            
            # Recommendations
            recommendations = []
            if critical_branches:
                recommendations.append(f"URGENTE: Redistribuir stock de {len(critical_branches)} sucursales críticas ({', '.join(b['branch_name'] for b in critical_branches[:3])})")
            if warning_branches:
                recommendations.append(f"Monitorear {len(warning_branches)} en alerta - planificar recepciones con cuidado")
            if no_limit_branches:
                recommendations.append(f"Configurar max_products en {len(no_limit_branches)} sucursales sin límite para control automático")
            if suggestions:
                recommendations.append(f"Ejecutar {len(suggestions)} transferencias sugeridas para balancear carga")
            
            if recommendations:
                h2("🎯 Recomendaciones")
                for i, rec in enumerate(recommendations, 1):
                    recommendation(f"{i}. {rec}")
            
            # Detail table
            h2("Capacidad por Sucursal")
            table(
                ["Sucursal", "SKUs actuales", "Máx.", "Uso %", "Estado", "Última actualización"],
                [[row["branch_name"], row["current_skus"], row["max_products"] if row["max_products"] else "Sin límite", 
                  f"{row['usage_percent']:.1f}%" if row["usage_percent"] is not None else "N/A", 
                  {"critical": "🔴 Crítica", "warning": "🟡 Alerta", "ok": "🟢 OK", "no_limit": "⚪ Sin límite"}.get(row["status"], row["status"]), 
                  row["last_inventory_update"]] for row in branches],
            )
            
            if suggestions:
                h2("Sugerencias de Transferencia")
                table(["Desde", "Hacia", "Motivo"], [[row["from_branch"], row["to_branch"], row["reason"]] for row in suggestions])

        elif report_type == "config_changes":
            changes = report_data.get("changes", [])
            summary = report_data.get("summary", {})
            total_changes = summary.get("total_changes", len(changes))
            by_type = summary.get("by_entity_type", {})
            
            # Executive Summary
            exec_lines = []
            exec_lines.append(f"Total cambios de configuración detectados: {total_changes:,}.")
            
            if by_type:
                top_type = max(by_type.items(), key=lambda x: x[1])
                exec_lines.append(f"Entidad más modificada: {top_type[0]} ({top_type[1]} cambios, {top_type[1]/total_changes*100:.1f}%).")
            
            # Recent high-impact changes
            sensitive_fields = ["price", "cost", "stock", "min_stock", "max_stock", "unit_price", "unit_cost", "is_active", "product_status"]
            sensitive_changes = [c for c in changes if any(f in c.get("field_name", "").lower() for f in sensitive_fields)]
            if sensitive_changes:
                exec_lines.append(f"⚠️  {len(sensitive_changes)} cambios en campos SENSIBLES (precio, stock, estado) - Auditar impacto.")
            
            if not changes:
                exec_lines.append("Sin cambios en el período.")
            
            executive_summary("\n".join(exec_lines))
            
            # Metrics
            h2("📊 Resumen Auditoría")
            metric_card("Total cambios", f"{total_changes:,}")
            metric_card("Entidades afectadas", str(len(by_type)))
            metric_card("Campos sensibles", str(len(sensitive_changes)), "warning" if sensitive_changes else "good")
            
            if by_type:
                h2("Por Tipo de Entidad")
                table(["Tipo", "Cambios", "%"], [[t, c, f"{c/total_changes*100:.1f}%"] for t, c in sorted(by_type.items(), key=lambda x: -x[1])])
            
            # Detail
            if changes:
                h2(f"Detalle (últimos 50 de {len(changes)})")
                table(["ID", "Entidad", "Campo", "Anterior", "Nuevo", "Por", "Fecha"], [[row["id"], row["entity_name"], row["field_name"], str(row["old_value"])[:50], str(row["new_value"])[:50], row["changed_by"], row["changed_at"]] for row in changes[:50]])
            
            # Recommendations
            if sensitive_changes:
                h2("🎯 Acciones Requeridas")
                recommendation("1. Revisar cada cambio en campos sensibles: validar autorización y impacto")
                recommendation("2. Verificar que cambios de precio/costo tengan aprobación gerencial documentada")
                recommendation("3. Confirmar que cambios de stock mínimos/máximos no rompan reposiciones automáticas")

        elif report_type == "batches":
            expiring_batches = report_data.get("expiring_batches", [])
            expired_batches = report_data.get("expired_batches", [])
            summary = report_data.get("summary", {})
            
            total_batches = summary.get("total_batches", len(expiring_batches) + len(expired_batches))
            expiring_soon = summary.get("expiring_soon", len(expiring_batches))
            expired = summary.get("expired", len(expired_batches))
            total_qty_at_risk = summary.get("total_quantity_at_risk", 0)
            
            urgent = [b for b in expiring_batches if b.get("days_until_expiry", 0) <= 7]
            warning = [b for b in expiring_batches if 7 < b.get("days_until_expiry", 0) <= 30]
            normal = [b for b in expiring_batches if b.get("days_until_expiry", 0) > 30]
            
            # Executive Summary
            exec_lines = []
            exec_lines.append(f"Análisis de lotes: {total_batches} totales | {expiring_soon} por vencer | {expired} expirados.")
            exec_lines.append(f"Cantidad en riesgo: {total_qty_at_risk:,} unidades.")
            
            if urgent:
                exec_lines.append(f"🔴 CRÍTICO: {len(urgent)} lotes vencen en ≤7 días - ACCIÓN INMEDIATA.")
            if warning:
                exec_lines.append(f"🟡 URGENTE: {len(warning)} lotes vencen en 8-30 días - Planear promo/transferencia.")
            if normal:
                exec_lines.append(f"🟢 MONITOR: {len(normal)} lotes vencen en >30 días - Seguimiento rutinario.")
            if expired:
                exec_lines.append(f"⚫ EXPIRADOS: {len(expired)} lotes ya vencidos - Disposición requerida.")
            if not expiring_batches and not expired_batches:
                exec_lines.append("✅ Sin lotes en riesgo de vencimiento.")
            
            executive_summary("\n".join(exec_lines))
            
            # Metrics
            h2("📊 Panel de Vencimientos")
            metric_card("Total lotes analizados", str(total_batches))
            metric_card("Próximos a vencer", str(expiring_soon), "critical" if urgent else ("warning" if warning else "good"))
            metric_card("  → ≤7 días (crítico)", str(len(urgent)), "critical" if urgent else "good")
            metric_card("  → 8-30 días (urgente)", str(len(warning)), "warning" if warning else "good")
            metric_card("  → >30 días (monitor)", str(len(normal)))
            metric_card("Ya expirados", str(expired), "critical" if expired else "good")
            metric_card("Unidades en riesgo", f"{total_qty_at_risk:,}")
            
            # Recommendations
            recommendations = []
            if urgent:
                recommendations.append(f"LANZAR PROMOCIONES FLASH para {len(urgent)} lotes críticos (descuento 20-50% o 2x1)")
                recommendations.append("Coordinar con marketing: comunicación urgente 'últimas unidades'")
            if warning:
                recommendations.append(f"Planear promociones estándar para {len(warning)} lotes en ventana 8-30 días")
            if expired:
                recommendations.append(f"Disponer {len(expired)} lotes expirados: donación, destrucción certificada o retorno a proveedor")
            if by_branch := report_data.get("by_branch", {}):
                high_risk_branches = [(b, v) for b, v in by_branch.items() if v.get("expiring", 0) + v.get("expired", 0) > 5]
                if high_risk_branches:
                    recommendations.append(f"Sucursales con alta concentración: {', '.join(b for b, v in high_risk_branches[:3])} - redistribuir a sucursales con mejor rotación")
            
            if recommendations:
                h2("🎯 Plan de Acción Inmediato")
                for i, rec in enumerate(recommendations, 1):
                    recommendation(f"{i}. {rec}")
            
            # Detail tables
            if expiring_batches:
                h2(f"Lotes Próximos a Vencer ({len(expiring_batches)})")
                table(
                    ["Lote", "Producto", "Sucursal", "Vence", "Días", "Cant.", "Urgencia", "Acción"],
                    [[row["batch_number"], row["product_name"], row["branch_name"], 
                      _fmt_date_iso_to_readable(row["expiration_date"]), row["days_until_expiry"], row["quantity"],
                      "🔴 ≤7d" if row["days_until_expiry"] <= 7 else ("🟡 8-30d" if row["days_until_expiry"] <= 30 else "🟢 >30d"),
                      "Promo FLASH" if row["days_until_expiry"] <= 7 else ("Promo/Transferencia" if row["days_until_expiry"] <= 30 else "Monitor")]
                     for row in expiring_batches[:30]],
                )
            
            if expired_batches:
                h2(f"⚫ Lotes Expirados ({len(expired_batches)})")
                table(
                    ["Lote", "Producto", "Sucursal", "Venció", "Días", "Cant.", "Disposición sugerida"],
                    [[row["batch_number"], row["product_name"], row["branch_name"], 
                      _fmt_date_iso_to_readable(row["expiration_date"]), row["days_until_expiry"], row["quantity"],
                      "Donar" if row["quantity"] < 10 else ("Destruir" if row["days_until_expiry"] < -30 else "Retornar proveedor")]
                     for row in expired_batches[:30]],
                )

        elif report_type == "kits":
            kits = report_data.get("kits", [])
            total_kits = len(kits)
            total_components = sum(len(k.get("components", [])) for k in kits)
            
            # Executive Summary
            exec_lines = []
            if total_kits > 0:
                exec_lines.append(f"Catálogo de Kits: {total_kits} kits definidos con {total_components} componentes totales.")
                
                # Stock status analysis
                out_of_stock = sum(1 for k in kits if k.get("stock_status") == "out")
                low_stock = sum(1 for k in kits if k.get("stock_status") == "low")
                ok_stock = sum(1 for k in kits if k.get("stock_status") == "ok")
                
                exec_lines.append(f"Estado de stock: {ok_stock} OK | {low_stock} BAJO | {out_of_stock} AGOTADO.")
                
                if out_of_stock > 0:
                    exec_lines.append(f"🔴 {out_of_stock} kits SIN STOCK - No se pueden armar/vender.")
                if low_stock > 0:
                    exec_lines.append(f"🟡 {low_stock} kits con stock LIMITADO - Priorizar reposición componentes.")
                
                # Component analysis
                all_components = []
                for k in kits:
                    for comp in k.get("components", []):
                        all_components.append(comp)
                
                if all_components:
                    critical_comps = [c for c in all_components if c.get("quantity_needed", 0) > 10]
                    exec_lines.append(f"Componentes críticos (necesidad >10/u): {len(critical_comps)}.")
            else:
                exec_lines.append("No hay kits definidos en el sistema.")
            
            executive_summary("\n".join(exec_lines))
            
            # Metrics
            h2("📊 Resumen de Kits")
            metric_card("Total kits", str(total_kits))
            if total_kits > 0:
                metric_card("Componentes únicos", str(total_components))
                metric_card("Stock OK", str(ok_stock), "good" if ok_stock == total_kits else ("warning" if low_stock > 0 else "critical"))
                metric_card("Stock Bajo", str(low_stock), "warning" if low_stock > 0 else "good")
                metric_card("Sin Stock", str(out_of_stock), "critical" if out_of_stock > 0 else "good")
            
            # Detail
            if kits:
                h2("Detalle por Kit")
                for kit in kits:
                    status_icon = {"ok": "🟢", "low": "🟡", "out": "🔴"}.get(kit.get("stock_status", "ok"), "⚪")
                    lines.append(f"\n  {status_icon} {kit['product_name']} ({kit['sku']}) - Stock kit: {kit.get('total_stock_kit', 0)} - Estado: {kit.get('stock_status', 'ok')}")
                    if kit.get("components"):
                        table(
                            ["Componente", "SKU", "Cant. requerida", "Stock disp."],
                            [[component["name"], component["sku"], component["quantity_needed"], "—"] 
                             for component in kit["components"]],
                        )

        elif report_type == "abc_analysis":
            s = report_data.get("summary", {})
            class_a = s.get("class_a_count", 0)
            class_b = s.get("class_b_count", 0)
            class_c = s.get("class_c_count", 0)
            dead_stock = s.get("dead_stock_count", 0)
            dead_stock_value = s.get("dead_stock_value", 0)
            total_products = class_a + class_b + class_c + dead_stock
            total_value = sum(item.get("total_value", 0) for item in report_data.get("class_a_products", []) + report_data.get("class_b_products", []) + report_data.get("class_c_products", []))
            
            # Executive Summary
            exec_lines = []
            if total_products > 0:
                exec_lines.append(f"Clasificación ABC completada: {total_products} productos analizados | Valor total inventario: ${total_value:,.2f}.")
                
                exec_lines.append(f"Clase A (Críticos): {class_a} productos ({class_a/total_products*100:.1f}%) → ~80% valor.")
                exec_lines.append(f"Clase B (Importantes): {class_b} productos ({class_b/total_products*100:.1f}%) → ~15% valor.")
                exec_lines.append(f"Clase C (Rutina): {class_c} productos ({class_c/total_products*100:.1f}%) → ~5% valor.")
                
                if dead_stock > 0:
                    exec_lines.append(f"🔴 DEAD STOCK: {dead_stock} productos ({dead_stock/total_products*100:.1f}%) SIN MOVIMIENTO → Valor atrapado: ${dead_stock_value:,.2f}.")
                    exec_lines.append("    → Acción: Liquidar, donar o retornar a proveedor.")
                
                # Pareto check
                class_a_value = sum(item.get("total_value", 0) for item in report_data.get("class_a_products", []))
                if total_value > 0:
                    a_pct = class_a_value / total_value * 100
                    exec_lines.append(f"Validación Pareto: Clase A = {a_pct:.1f}% del valor (esperado ~80%).")
            else:
                exec_lines.append("Sin productos para clasificar.")
            
            executive_summary("\n".join(exec_lines))
            
            # Metrics
            h2("📊 Scorecard ABC")
            metric_card("Total productos", str(total_products))
            metric_card("Clase A (Críticos)", f"{class_a} ({class_a/total_products*100:.1f}%)", "critical" if class_a > 0 else "good")
            metric_card("Clase B (Importantes)", f"{class_b} ({class_b/total_products*100:.1f}%)", "warning" if class_b > 0 else "good")
            metric_card("Clase C (Rutina)", f"{class_c} ({class_c/total_products*100:.1f}%)", "good")
            metric_card("Dead Stock", f"{dead_stock} ({dead_stock/total_products*100:.1f}%)", "critical" if dead_stock > 0 else "good")
            metric_card("Valor Dead Stock", f"${dead_stock_value:,.2f}", "critical" if dead_stock_value > 0 else "good")
            metric_card("Valor Total Inventario", f"${total_value:,.2f}")
            
            # Recommendations
            recommendations = []
            if class_a > 0:
                recommendations.append("Clase A: Configurar alertas automáticas, conteo cíclico semanal, stock de seguridad alto")
            if dead_stock > 0:
                recommendations.append(f"Dead Stock ({dead_stock} items): Campaña liquidación 50-70% dto, donación fiscal, o retorno proveedor")
                recommendations.append(f"Valor recuperable estimado: ${dead_stock_value * 0.3:,.2f} (30% valor libros) vs ${dead_stock_value:,.2f} costo oportunidad")
            if class_c > total_products * 0.6:
                recommendations.append("Muchos productos Clase C: Revisar catálogo - ¿eliminar SKUs zombis?")
            if total_products > 1000:
                recommendations.append("Catálogo >1000 SKUs: Implementar segmentación automática mensual")
            
            if recommendations:
                h2("🎯 Estrategia por Clase")
                for i, rec in enumerate(recommendations, 1):
                    recommendation(f"{i}. {rec}")
            
            # Detail tables
            if report_data.get("class_a_products"):
                h2("🔴 Clase A - Monitoreo CONTINUO (Top 20)")
                table(["Producto", "SKU", "Valor Total", "% Total", "Acción"], [[item["product_name"], item["sku"], f"${item['total_value']:,.2f}", item["percent_of_total"], "Alertas + Conteo semanal + Stock seguro ALTO"] for item in report_data["class_a_products"][:20]])
            
            if report_data.get("class_b_products"):
                h2("🟡 Clase B - Revisión PERIÓDICA (Top 20)")
                table(["Producto", "SKU", "Valor Total", "% Total"], [[item["product_name"], item["sku"], f"${item['total_value']:,.2f}", item["percent_of_total"]] for item in report_data["class_b_products"][:20]])
            
            if report_data.get("class_c_products"):
                h2("🟢 Clase C - Control BÁSICO (Top 20)")
                table(["Producto", "SKU", "Valor Total", "% Total"], [[item["product_name"], item["sku"], f"${item['total_value']:,.2f}", item["percent_of_total"]] for item in report_data["class_c_products"][:20]])
            
            if report_data.get("dead_stock"):
                h2("⚫ DEAD STOCK - Plan de Liquidación")
                table(["Producto", "SKU", "Días sin mov.", "Stock actual", "Valor atrapado", "Acción recomendada"], [[item["product_name"], item["sku"], item["days_without_movement"], item["current_stock"], f"${item['total_value']:,.2f}" if 'total_value' in item else "$—", "Liquidar 70% dto / Donar / Devolver"] for item in report_data["dead_stock"][:20]])

        return "\n".join(lines)

    # ─────────────────────────────────────────────────────────────────────────
    # EXP 7 · REPORTES GUARDADOS
    # ─────────────────────────────────────────────────────────────────────────

    def save_report_config(
        self,
        name: str,
        report_type: str,
        parameters: Dict[str, Any],
        user_id: Optional[int] = None,
    ) -> Dict[str, Any]:
        """Persist a named report configuration. Returns the saved record."""
        from models.saved_report import SavedReport

        record = SavedReport(
            name=name,
            report_type=report_type,
            parameters=json.dumps(parameters, ensure_ascii=False),
            created_by=user_id,
        )
        self.db.add(record)
        self.db.commit()
        self.db.refresh(record)
        logger.info("Saved report config id=%d name='%s'", record.id, record.name)
        return record.to_dict()

    def get_saved_reports(
        self, user_id: Optional[int] = None
    ) -> List[Dict[str, Any]]:
        """Return all saved report configurations, optionally filtered by user."""
        from models.saved_report import SavedReport

        query = self.db.query(SavedReport)
        if user_id is not None:
            query = query.filter(SavedReport.created_by == user_id)
        records = query.order_by(SavedReport.created_at.desc()).all()
        return [r.to_dict() for r in records]

    def get_report_config(self, config_id: int) -> Optional[Dict[str, Any]]:
        """Return a single saved report config by id, or None."""
        from models.saved_report import SavedReport

        record = self.db.query(SavedReport).filter(SavedReport.id == config_id).first()
        if not record:
            return None
        data = record.to_dict()
        # Deserialise parameters for convenience
        try:
            data["parameters_dict"] = json.loads(record.parameters or "{}")
        except (json.JSONDecodeError, TypeError):
            data["parameters_dict"] = {}
        return data

    def delete_report_config(self, config_id: int) -> bool:
        """Delete a saved report config. Returns True on success."""
        from models.saved_report import SavedReport

        record = self.db.query(SavedReport).filter(SavedReport.id == config_id).first()
        if not record:
            return False
        self.db.delete(record)
        self.db.commit()
        logger.info("Deleted saved report config id=%d", config_id)
        return True

    # ─────────────────────────────────────────────────────────────────────────
    # EXP 8 · COMPARACIÓN DE PERÍODOS
    # ─────────────────────────────────────────────────────────────────────────

    def generate_comparison_report(
        self,
        report_type: str,
        period1_from: datetime,
        period1_to: datetime,
        period2_from: datetime,
        period2_to: datetime,
        branch_id: Optional[int] = None,
    ) -> Dict[str, Any]:
        """
        Generate the same report for two date ranges and compute deltas.

        report_type: 'inventory' | 'movements' | 'discrepancies' | 'kpis'
        """
        GENERATORS = {
            "inventory": self.generate_inventory_report,
            "movements": self.generate_movement_report,
            "discrepancies": self.generate_discrepancy_report,
            "kpis": self.generate_kpi_report,
        }
        gen = GENERATORS.get(report_type)
        if gen is None:
            raise ValueError(
                f"report_type '{report_type}' no soporta comparación. "
                f"Usa: {list(GENERATORS.keys())}"
            )

        p1 = gen(branch_id=branch_id, date_from=period1_from, date_to=period1_to)
        p2 = gen(branch_id=branch_id, date_from=period2_from, date_to=period2_to)

        # Build numeric differences for scalar top-level fields
        diffs: Dict[str, Any] = {}
        numeric_keys = {
            k for k in p1
            if isinstance(p1[k], (int, float))
            and k not in ("branch_id",)
        }
        for k in numeric_keys:
            v1, v2 = p1.get(k, 0) or 0, p2.get(k, 0) or 0
            diffs[k] = {
                "period1": v1,
                "period2": v2,
                "delta": v2 - v1,
                "delta_pct": round(((v2 - v1) / v1 * 100) if v1 else 0, 2),
            }

        return {
            "generated_at": _fmt_date(datetime.utcnow()),
            "report_type": report_type,
            "branch_id": branch_id,
            "period1": {
                "from": _fmt_date(period1_from),
                "to": _fmt_date(period1_to),
                "data": p1,
            },
            "period2": {
                "from": _fmt_date(period2_from),
                "to": _fmt_date(period2_to),
                "data": p2,
            },
            "differences": diffs,
        }

    # ─────────────────────────────────────────────────────────────────────────
    # EXP 9 · TOP PRODUCTOS
    # ─────────────────────────────────────────────────────────────────────────

    def generate_top_products_report(
        self,
        metric: str = "most_moved",
        limit: int = 10,
        branch_id: Optional[int] = None,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
    ) -> Dict[str, Any]:
        """
        Return the top *limit* products ranked by *metric*.

        metric values:
            most_moved       – most validated movements (qty)
            least_moved      – fewest validated movements
            most_stock       – highest digital stock
            lowest_stock     – lowest digital stock (> 0)
            most_discrepancies – largest absolute discrepancy
        """
        from models.movement import Movement, MovementState
        from models.inventory import Inventory
        from models.product import Product
        from models.branch import Branch

        products: List[Dict[str, Any]] = []

        if metric in ("most_moved", "least_moved"):
            query = (
                self.db.query(
                    Product.id,
                    Product.name,
                    Product.sku,
                    func.count(Movement.id).label("move_count"),
                    func.sum(Movement.quantity).label("total_qty"),
                )
                .join(Movement, Movement.product_id == Product.id)
                .filter(
                    Movement.state == MovementState.VALIDADO.value,
                    Movement.is_cancelled == False,
                )
            )
            if branch_id:
                query = query.filter(Movement.branch_id == branch_id)
            if date_from:
                query = query.filter(Movement.created_at >= date_from)
            if date_to:
                query = query.filter(Movement.created_at <= date_to)
            query = query.group_by(Product.id, Product.name, Product.sku)
            asc = metric == "least_moved"
            order_col = func.sum(Movement.quantity)
            query = query.order_by(order_col if asc else order_col.desc())
            query = query.limit(limit)

            for rank, row in enumerate(query.all(), start=1):
                products.append({
                    "rank": rank,
                    "product_id": row.id,
                    "product": row.name,
                    "sku": row.sku,
                    "move_count": row.move_count,
                    "total_quantity_moved": row.total_qty or 0,
                })

        elif metric in ("most_stock", "lowest_stock"):
            query = (
                self.db.query(
                    Product.id,
                    Product.name,
                    Product.sku,
                    func.sum(Inventory.digital_stock).label("total_stock"),
                )
                .join(Inventory, Inventory.product_id == Product.id)
                .join(Branch, Branch.id == Inventory.branch_id)
                .filter(
                    Inventory.is_active == True,
                    Product.is_active == True,
                    Branch.is_active == True,
                )
            )
            if branch_id:
                query = query.filter(Inventory.branch_id == branch_id)
            query = query.group_by(Product.id, Product.name, Product.sku)
            if metric == "lowest_stock":
                query = query.filter(
                    func.sum(Inventory.digital_stock) > 0
                ).order_by(func.sum(Inventory.digital_stock))
            else:
                query = query.order_by(func.sum(Inventory.digital_stock).desc())
            query = query.limit(limit)

            for rank, row in enumerate(query.all(), start=1):
                products.append({
                    "rank": rank,
                    "product_id": row.id,
                    "product": row.name,
                    "sku": row.sku,
                    "total_stock": row.total_stock or 0,
                })

        elif metric == "most_discrepancies":
            query = (
                self.db.query(
                    Product.id,
                    Product.name,
                    Product.sku,
                    func.sum(
                        func.abs(Inventory.physical_stock - Inventory.digital_stock)
                    ).label("total_diff"),
                )
                .join(Inventory, Inventory.product_id == Product.id)
                .join(Branch, Branch.id == Inventory.branch_id)
                .filter(
                    Inventory.physical_stock != Inventory.digital_stock,
                    Inventory.is_active == True,
                    Product.is_active == True,
                    Branch.is_active == True,
                )
            )
            if branch_id:
                query = query.filter(Inventory.branch_id == branch_id)
            query = (
                query.group_by(Product.id, Product.name, Product.sku)
                .order_by(func.sum(
                    func.abs(Inventory.physical_stock - Inventory.digital_stock)
                ).desc())
                .limit(limit)
            )

            for rank, row in enumerate(query.all(), start=1):
                products.append({
                    "rank": rank,
                    "product_id": row.id,
                    "product": row.name,
                    "sku": row.sku,
                    "total_discrepancy": row.total_diff or 0,
                })
        else:
            raise ValueError(
                f"Métrica desconocida: '{metric}'. Usa: most_moved, least_moved, "
                "most_stock, lowest_stock, most_discrepancies"
            )

        return {
            "generated_at": _fmt_date(datetime.utcnow()),
            "metric": metric,
            "limit": limit,
            "filters": {
                "branch_id": branch_id,
                "date_from": _fmt_date(date_from),
                "date_to": _fmt_date(date_to),
            },
            "products": products,
        }

    # ─────────────────────────────────────────────────────────────────────────
    # EXP 10 · EFICIENCIA POR SUCURSAL
    # ─────────────────────────────────────────────────────────────────────────

    def generate_branch_efficiency_report(
        self,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
    ) -> Dict[str, Any]:
        """
        Calculate per-branch operational metrics:
          - total_movements, rejection_rate, avg_validation_days,
            transfers_sent, transfers_received, discrepancy_rate
        """
        from models.movement import Movement, MovementState, MovementType
        from models.inventory import Inventory
        from models.branch import Branch
        from models.product import Product

        branches = (
            self.db.query(Branch)
            .filter(Branch.is_active == True)
            .order_by(Branch.name)
            .all()
        )

        branch_data: List[Dict[str, Any]] = []

        for branch in branches:
            q = self.db.query(Movement).filter(Movement.branch_id == branch.id)
            if date_from:
                q = q.filter(Movement.created_at >= date_from)
            if date_to:
                q = q.filter(Movement.created_at <= date_to)
            all_mvs = q.all()

            total = len(all_mvs)
            rejected = sum(1 for m in all_mvs if m.state == MovementState.RECHAZADO.value)
            rejection_rate = round((rejected / total * 100) if total else 0, 2)

            # Avg validation days (validated movements only)
            validated = [
                m for m in all_mvs
                if m.state == MovementState.VALIDADO.value
                and m.validated_at and m.created_at
            ]
            if validated:
                total_days = sum(
                    (m.validated_at - m.created_at).total_seconds() / 86400
                    for m in validated
                )
                avg_validation_days = round(total_days / len(validated), 2)
            else:
                avg_validation_days = None

            transfers_sent = sum(
                1 for m in all_mvs
                if m.movement_type == MovementType.TRANSFERENCIA.value
            )
            transfers_received = (
                self.db.query(Movement)
                .filter(
                    Movement.destination_branch_id == branch.id,
                    Movement.movement_type == MovementType.TRANSFERENCIA.value,
                )
            )
            if date_from:
                transfers_received = transfers_received.filter(
                    Movement.created_at >= date_from
                )
            if date_to:
                transfers_received = transfers_received.filter(
                    Movement.created_at <= date_to
                )
            transfers_received_count = transfers_received.count()

            # Discrepancy rate = items with discrepancy / total active items
            total_items = (
                self.db.query(Inventory)
                .join(Product)
                .filter(
                    Inventory.branch_id == branch.id,
                    Inventory.is_active == True,
                    Product.is_active == True,
                )
                .count()
            )
            disc_items = (
                self.db.query(Inventory)
                .join(Product)
                .filter(
                    Inventory.branch_id == branch.id,
                    Inventory.is_active == True,
                    Product.is_active == True,
                    Inventory.physical_stock != Inventory.digital_stock,
                )
                .count()
            )
            discrepancy_rate = round(
                (disc_items / total_items * 100) if total_items else 0, 2
            )

            branch_data.append({
                "branch_id": branch.id,
                "branch": branch.name,
                "total_movements": total,
                "rejection_rate": rejection_rate,
                "avg_validation_days": avg_validation_days,
                "transfers_sent": transfers_sent,
                "transfers_received": transfers_received_count,
                "discrepancy_rate": discrepancy_rate,
            })

        return {
            "generated_at": _fmt_date(datetime.utcnow()),
            "filters": {
                "date_from": _fmt_date(date_from),
                "date_to": _fmt_date(date_to),
            },
            "branches": branch_data,
        }

    # ─────────────────────────────────────────────────────────────────────────
    # EXP 11 · VALOR DE INVENTARIO
    # ─────────────────────────────────────────────────────────────────────────

    def generate_inventory_value_report(
        self,
        branch_id: Optional[int] = None,
    ) -> Dict[str, Any]:
        """
        Calculate the monetary value of current stock.

        Unit price resolution order:
          1. Inventory.unit_cost  (branch-specific cost)
          2. Product.unit_price   (global fallback)

        Items without any price are included with value=0 and flagged.
        """
        from models.inventory import Inventory
        from models.product import Product
        from models.branch import Branch

        query = (
            self.db.query(Inventory)
            .join(Product)
            .join(Branch)
            .filter(
                Inventory.is_active == True,
                Product.is_active == True,
                Branch.is_active == True,
            )
        )
        if branch_id:
            query = query.filter(Inventory.branch_id == branch_id)

        items = query.order_by(Branch.name, Product.name).all()

        total_value = 0.0
        by_branch: Dict[str, Any] = {}
        item_rows: List[Dict[str, Any]] = []
        no_price_count = 0

        for item in items:
            unit_price = item.unit_cost or item.product.unit_price
            has_price = unit_price is not None
            value = round((item.digital_stock * unit_price) if has_price else 0.0, 2)
            if not has_price:
                no_price_count += 1

            total_value += value

            branch_name = item.branch.name
            if branch_name not in by_branch:
                by_branch[branch_name] = {"items": 0, "value": 0.0}
            by_branch[branch_name]["items"] += 1
            by_branch[branch_name]["value"] = round(
                by_branch[branch_name]["value"] + value, 2
            )

            item_rows.append({
                "product_id": item.product_id,
                "product": item.product.name,
                "sku": item.product.sku,
                "branch": branch_name,
                "stock": item.digital_stock,
                "unit_price": unit_price,
                "value": value,
                "no_price": not has_price,
            })

        # Sort by value descending
        item_rows.sort(key=lambda x: x["value"], reverse=True)

        return {
            "generated_at": _fmt_date(datetime.utcnow()),
            "filters": {"branch_id": branch_id},
            "total_value": round(total_value, 2),
            "total_items": len(items),
            "no_price_count": no_price_count,
            "by_branch": by_branch,
            "items": item_rows,
        }

    # ─────────────────────────────────────────────────────────────────────────
    # EXP 12 · REPORTE DE TRANSFERENCIAS
    # ─────────────────────────────────────────────────────────────────────────

    def generate_transfer_report(
        self,
        branch_id: Optional[int] = None,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
    ) -> Dict[str, Any]:
        """
        Detailed report of inter-branch transfers.

        Includes: per-transfer detail, totals by branch pair,
        and a list of transfers pending reception.
        """
        from models.movement import Movement, MovementType
        from models.product import Product
        from models.branch import Branch

        query = (
            self.db.query(Movement)
            .filter(Movement.movement_type == MovementType.TRANSFERENCIA.value)
        )
        if branch_id:
            query = query.filter(
                or_(
                    Movement.branch_id == branch_id,
                    Movement.destination_branch_id == branch_id,
                )
            )
        if date_from:
            query = query.filter(Movement.created_at >= date_from)
        if date_to:
            query = query.filter(Movement.created_at <= date_to)

        transfers = query.order_by(Movement.created_at.desc()).all()

        # Pre-load branch names
        branch_cache: Dict[int, str] = {}

        def branch_name(bid: Optional[int]) -> str:
            if bid is None:
                return "—"
            if bid not in branch_cache:
                b = self.db.query(Branch).filter(Branch.id == bid).first()
                branch_cache[bid] = b.name if b else str(bid)
            return branch_cache[bid]

        # Pre-load product names
        product_cache: Dict[int, str] = {}

        def product_name(pid: int) -> str:
            if pid not in product_cache:
                p = self.db.query(Product).filter(Product.id == pid).first()
                product_cache[pid] = p.name if p else str(pid)
            return product_cache[pid]

        items: List[Dict[str, Any]] = []
        by_pair: Dict[str, Any] = {}
        pending_reception = 0

        for t in transfers:
            origin = branch_name(t.branch_id)
            destination = branch_name(t.destination_branch_id)
            pair_key = f"{origin} → {destination}"

            if pair_key not in by_pair:
                by_pair[pair_key] = {"count": 0, "total_quantity": 0}
            by_pair[pair_key]["count"] += 1
            by_pair[pair_key]["total_quantity"] += t.quantity

            if not t.is_received and not t.is_cancelled:
                pending_reception += 1

            items.append({
                "id": t.id,
                "product": product_name(t.product_id),
                "origin_branch": origin,
                "destination_branch": destination,
                "quantity": t.quantity,
                "state": t.state,
                "is_received": t.is_received,
                "is_cancelled": t.is_cancelled,
                "created_at": _fmt_date(t.created_at),
                "received_at": _fmt_date(t.received_at),
                "received_notes": t.received_notes,
            })

        return {
            "generated_at": _fmt_date(datetime.utcnow()),
            "filters": {
                "branch_id": branch_id,
                "date_from": _fmt_date(date_from),
                "date_to": _fmt_date(date_to),
            },
            "total_transfers": len(transfers),
            "pending_reception": pending_reception,
            "by_branch_pair": by_pair,
            "items": items,
        }

    # ─────────────────────────────────────────────────────────────────────────
    # EXP 13 · TENDENCIAS
    # ─────────────────────────────────────────────────────────────────────────

    # ─────────────────────────────────────────────────────────────────────────
    # HISTORIAL · AUDITORÍA DE HISTORIAL
    # ─────────────────────────────────────────────────────────────────────────

    def generate_history_audit_report(
        self,
        entity_type: Optional[str] = None,
        branch_id: Optional[int] = None,
        user_id: Optional[int] = None,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
        event_type: Optional[str] = None,
        limit: int = 500,
    ) -> Dict[str, Any]:
        """
        Auditoría completa del historial de eventos del sistema.

        El historial solo provee los datos en bruto; este método los agrega,
        cuenta y estructura para que Reports pueda exportarlos/formatearlos.

        Parámetros
        ──────────
        entity_type : filtrar por tipo de entidad (product, branch, movement,
                      inventory, alert, user, system — None = todos)
        branch_id   : filtrar entradas que mencionen esta sucursal en details
                      (usa la misma estrategia LIKE del historial, Expansión 1)
        user_id     : filtrar por usuario que realizó la acción
        date_from / date_to : rango de fechas
        event_type  : filtrar por nombre exacto de evento (ej. "product.updated")
        limit       : máximo de entradas a incluir en el detalle (def. 500)

        Retorno
        ───────
        {
          "generated_at": ...,
          "filters": {...},
          "total_entries": int,
          "by_event_type": {event_type: count, ...},
          "by_entity_type": {entity_type: count, ...},
          "by_user": {user_name: count, ...},
          "by_date": {YYYY-MM-DD: count, ...},       ← actividad diaria
          "top_entities": [{entity_type, entity_name, event_count}, ...],
          "items": [{id, fecha, evento, tipo, entidad, usuario, acción}, ...]
        }
        """
        from modules.history.service import HistoryService

        history_svc = HistoryService(self.db)

        # ── Obtener entradas con enriquecimiento de nombres (Exp 3) ──────────
        result = history_svc.list_history(
            limit=limit,
            event_type=event_type,
            entity_type=entity_type,
            user_id=user_id,
            branch_id=branch_id,
            date_from=date_from,
            date_to=date_to,
            enrich=True,          # resuelve user_name y entity_name
        )
        entries = result["entries"]
        total = result["total"]

        # ── Agregaciones ─────────────────────────────────────────────────────
        by_event_type: Dict[str, int] = {}
        by_entity_type: Dict[str, int] = {}
        by_user: Dict[str, int] = {}
        by_date: Dict[str, int] = {}
        entity_counts: Dict[str, int] = {}   # "entity_type|entity_name" → count

        items: List[Dict[str, Any]] = []

        for e in entries:
            ev = e.get("event_type") or "—"
            ent_type = e.get("entity_type") or "—"
            user_name = e.get("user_name") or "—"
            entity_name = e.get("entity_name") or "—"
            created_raw = e.get("created_at") or ""
            date_key = created_raw[:10] if created_raw else "—"

            by_event_type[ev] = by_event_type.get(ev, 0) + 1
            by_entity_type[ent_type] = by_entity_type.get(ent_type, 0) + 1
            by_user[user_name] = by_user.get(user_name, 0) + 1
            by_date[date_key] = by_date.get(date_key, 0) + 1

            ec_key = f"{ent_type}|{entity_name}"
            entity_counts[ec_key] = entity_counts.get(ec_key, 0) + 1

            items.append({
                "id":          e.get("id"),
                "fecha":       created_raw[:19].replace("T", " "),
                "evento":      ev,
                "tipo":        ent_type,
                "entidad":     entity_name,
                "usuario":     user_name,
                "accion":      e.get("action") or "—",
                "eliminado":   e.get("entity_deleted", False),
            })

        # ── Top entidades más activas (máx. 20) ───────────────────────────────
        top_entities = sorted(
            [
                {
                    "entity_type": k.split("|")[0],
                    "entity_name": k.split("|")[1],
                    "event_count": v,
                }
                for k, v in entity_counts.items()
            ],
            key=lambda x: x["event_count"],
            reverse=True,
        )[:20]

        # ── Actividad diaria ordenada ─────────────────────────────────────────
        by_date_sorted = dict(sorted(by_date.items()))

        return {
            "generated_at": _fmt_date(datetime.utcnow()),
            "filters": {
                "entity_type": entity_type,
                "event_type":  event_type,
                "branch_id":   branch_id,
                "user_id":     user_id,
                "date_from":   _fmt_date(date_from),
                "date_to":     _fmt_date(date_to),
                "limit":       limit,
            },
            "total_entries":   total,
            "returned_entries": len(items),
            "by_event_type":   dict(sorted(by_event_type.items(), key=lambda x: x[1], reverse=True)),
            "by_entity_type":  dict(sorted(by_entity_type.items(), key=lambda x: x[1], reverse=True)),
            "by_user":         dict(sorted(by_user.items(), key=lambda x: x[1], reverse=True)),
            "by_date":         by_date_sorted,
            "top_entities":    top_entities,
            "items":           items,
        }

    def generate_trend_report(
        self,
        metric: str = "movimientos_total",
        period_days: int = 7,
        periods_back: int = 8,
        branch_id: Optional[int] = None,
    ) -> Dict[str, Any]:
        """
        Generate time-series trend data for the last *periods_back* windows
        of *period_days* each.

        metric values:
            movimientos_total   – total validated movements per window
            stock_total         – digital stock snapshot (end of window)
            discrepancias_count – items with discrepancy per window
        """
        from models.movement import Movement, MovementState
        from models.inventory import Inventory
        from models.product import Product
        from models.branch import Branch

        now = datetime.utcnow()
        data_points: List[Dict[str, Any]] = []

        for i in range(periods_back - 1, -1, -1):
            window_end = now - timedelta(days=i * period_days)
            window_start = window_end - timedelta(days=period_days)
            period_label = window_start.strftime("%Y-%m-%d")

            if metric == "movimientos_total":
                q = self.db.query(func.count(Movement.id)).filter(
                    Movement.state == MovementState.VALIDADO.value,
                    Movement.is_cancelled == False,
                    Movement.created_at >= window_start,
                    Movement.created_at < window_end,
                )
                if branch_id:
                    q = q.filter(Movement.branch_id == branch_id)
                value = q.scalar() or 0

            elif metric == "stock_total":
                # Stock snapshot: use items updated before window_end
                q = self.db.query(func.sum(Inventory.digital_stock)).join(
                    Product
                ).join(Branch).filter(
                    Inventory.is_active == True,
                    Product.is_active == True,
                    Branch.is_active == True,
                )
                if branch_id:
                    q = q.filter(Inventory.branch_id == branch_id)
                value = q.scalar() or 0  # current snapshot (no time travel on stock)

            elif metric == "discrepancias_count":
                q = self.db.query(func.count(Inventory.id)).join(
                    Product
                ).join(Branch).filter(
                    Inventory.is_active == True,
                    Product.is_active == True,
                    Branch.is_active == True,
                    Inventory.physical_stock != Inventory.digital_stock,
                )
                if branch_id:
                    q = q.filter(Inventory.branch_id == branch_id)
                value = q.scalar() or 0
            else:
                raise ValueError(
                    f"Métrica desconocida: '{metric}'. Usa: movimientos_total, "
                    "stock_total, discrepancias_count"
                )

            data_points.append({
                "period": period_label,
                "window_start": _fmt_date(window_start),
                "window_end": _fmt_date(window_end),
                "value": value,
            })

        # Simple trend direction
        if len(data_points) >= 2:
            first_val = data_points[0]["value"]
            last_val = data_points[-1]["value"]
            if last_val > first_val:
                trend = "ascendente"
            elif last_val < first_val:
                trend = "descendente"
            else:
                trend = "estable"
        else:
            trend = "sin datos"

        return {
            "generated_at": _fmt_date(datetime.utcnow()),
            "metric": metric,
            "period_days": period_days,
            "periods_count": periods_back,
            "trend": trend,
            "filters": {"branch_id": branch_id},
            "data_points": data_points,
        }
